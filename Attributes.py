import pandas as pd
import os
from dotenv import load_dotenv
import numpy as np
import re
import openpyxl
from tqdm import tqdm  # Import tqdm

load_dotenv()

# 1 load GS1 & Maxeda datamodels
# 2 perform changes on picklistvalues (inherent with deletions) so that picklist-attributes can be delete later on
# 3 additions, deletions and changes for S7, S8, S23 (S7 can also change)
# 4 additions to picklistvalues for new picklists and attributes changed to picklist) 

gs1_file_path = os.getenv('path_datamodel_GS1')
datamodel_file_path = os.getenv('path_datamodel_maxeda')


###################
## GS1 datamodel
###################
print('### Read GS1 datamodel ###')

print('## Establish active brick ##')
# Read the 'Bricks' to select only the attributes from active Bricks
gs1_df_bricks = pd.read_excel(gs1_file_path, sheet_name='Bricks', skiprows=3, dtype=str)
gs1_active_bricks_set = set(gs1_df_bricks[gs1_df_bricks['Brick activated'] == 'Yes']['Brick Code'].dropna())

print('## Establish attributes from active brick ##')
# Read the Attributes per Brick sheet from the GS1 file to be able to address the attributes from active bricks
gs1_df_attributes_brick = pd.read_excel(gs1_file_path, sheet_name='Data for Attributes per Brick', skiprows=3, dtype=str)
# Filter gs1_df_attributes_brick for only those rows where the 'Brick' column's values are in gs1_active_bricks_set
gs1_df_attributes_brick_active = gs1_df_attributes_brick[gs1_df_attributes_brick['Brick'].isin(gs1_active_bricks_set)]

# Create a set of the 'FieldID' values from the filtered DataFrame
gs1_attributes_GS1ID_set = set(gs1_df_attributes_brick_active['FieldID'].dropna())

print('## Read attribute metadata ##')
# Read metadata for attributes
gs1_df_attributes = pd.read_excel(gs1_file_path, sheet_name='Fielddefinitions', skiprows=3, dtype=str)

print('## Read picklists ##')
# Read picklists
gs1_df_picklists = pd.read_excel(gs1_file_path, sheet_name='Picklists', skiprows=3, dtype=str)


###################
## Maxeda datamodel
###################

# Extract attribute code after "id" (case-insenstive)
def extract_attribute_code(definition):
    # Convert to string in case the input is NaN or any other non-string data
    definition = str(definition)
    pattern = re.compile(r"GS1 Field[_ ]ID (\S+)|GS1 FieldID(\S+)", re.IGNORECASE)
    match = pattern.search(definition)
    if match:
        return match.group(1).strip()  # Extract and return the part after 'id '
    return ''  # Return an empty string if no match is found

print(f'### Read Maxeda datamodel ###')
print(f'## Read S7 ##')
# Read the 'S7 - Attribute' sheet from the Datamodel file
maxeda_s7_full_df = pd.read_excel(datamodel_file_path, sheet_name='S7 - Attribute')

# Select scope
maxeda_s7_df_scope = maxeda_s7_full_df[maxeda_s7_full_df['Attribute Type'].isin(['Category', 'Common'])].copy()
# maxeda_s7_df_scope = maxeda_s7_full_df[maxeda_s7_full_df['Attribute Type'].isin(['Category'])].copy()

# Extract Attribute codes
maxeda_s7_df_scope['Attribute code'] = maxeda_s7_df_scope['Definition'].apply(extract_attribute_code)

# Exclude maxeda-attributes
maxeda_s7_df_scope = maxeda_s7_df_scope[~maxeda_s7_df_scope['Attribute code'].str.startswith("M")]

# Convert the 'Precision' column to string type
maxeda_s7_df_scope['Precision'] = maxeda_s7_df_scope['Precision'].astype(str)
# Use string manipulation to remove trailing '.0' from the string representation
maxeda_s7_df_scope['Precision'] = maxeda_s7_df_scope['Precision'].str.replace(r'\.0$', '', regex=True)
# Replace 'nan' strings with empty strings
maxeda_s7_df_scope['Precision'] = maxeda_s7_df_scope['Precision'].replace('nan', '')


print(f'## Read S8 ##')
# S8 - Attribute - Locale
maxeda_s8_full_df = pd.read_excel(datamodel_file_path, sheet_name='S8 - Attribute - Locale')

# Select scope
maxeda_s8_df_scope = maxeda_s8_full_df[maxeda_s8_full_df['Attribute Path'].str.contains('Category Specific Attributes//|OneWS_.*OneWS_', na=False)].copy()

# Convert the 'ID' column of maxeda_s8_df to a set
maxeda_s8_IDs_set = set(maxeda_s8_df_scope['ID'])


maxeda_s8_df_scope['Attribute code'] = maxeda_s8_df_scope['Definition'].apply(extract_attribute_code)
# Correct for the fact that not every attributeID is formatted with a dot on the second position
maxeda_s8_df_scope['Attribute code'] = maxeda_s8_df_scope['Attribute code'].apply(
    lambda x: x[0] + '.' + x[1:] if len(x) > 1 and x[1] != '.' else x
)

# Exclude maxeda-attributes
maxeda_s8_df_scope = maxeda_s8_df_scope[~maxeda_s8_df_scope['Attribute code'].str.startswith("M")]

print(f'## Read S23 - Lookup Model ##')
maxeda_s23_df = pd.read_excel(datamodel_file_path, sheet_name='S23 - Lookup Model')


####################
## LookupData changes & delete
####################

print(f'### Lookup tables ###')
maxeda_lookuptables_s7_set = set(maxeda_s7_df_scope['LookUp Table Name'].replace('', np.nan).dropna())

valid_combinations = set(zip(gs1_df_picklists['Picklist ID'], gs1_df_picklists['Code value']))

def read_excel_with_openpyxl(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook[sheet_name]
    data = []

    for row in sheet.iter_rows(values_only=True):
        data.append([str(cell).strip() if cell is not None else "" for cell in row])

    # Check if the required headers are present and if the second column has at least one non-empty value
    if data and len(data[0]) >= 2 and data[0][0] == "Id" and "Code" in data[0][1]:
        df = pd.DataFrame(data[1:], columns=data[0])
        return df
    return None  # Return None if conditions are not met

def read_all_excel_files(directory_path):
    counter = 0
    all_data_frames = []
    
    for filename in tqdm(os.listdir(directory_path), desc="Processing files"):
        if filename.endswith('.xlsx'):
            counter += 1
            if counter < 1000:
                # print(f"counter LookupData files: {counter}")
                file_path = os.path.join(directory_path, filename)
                workbook = openpyxl.load_workbook(file_path, data_only=True)

                # Iterate over each sheet except the first one
                for sheet_name in workbook.sheetnames[1:]:
                    if sheet_name in maxeda_lookuptables_s7_set: # only consider the sheets that are LookupTables for s7-categoeie variables

                        disregarded_entries = []
                        added_entries = []

                        df = read_excel_with_openpyxl(file_path, sheet_name) # read the Excel-sheet
                        if df is not None:
                            # Determine GS1 field ID
                            attribute_codes = maxeda_s7_df_scope.loc[maxeda_s7_df_scope['LookUp Table Name'] == sheet_name, 'Attribute code'] # Lookup the GS1-attribute code for this LookupData
                            attribute_code = attribute_codes.iloc[0] 

                            if attribute_code == '':
                                continue                       

                            # Next, determine the GS1 picklist id
                            # print("# Lookup picklist id #")
                            picklist_ids = gs1_df_attributes.loc[gs1_df_attributes['FieldID'] == attribute_code, 'Picklist ID']
                            if picklist_ids.empty:
                                # print(f"No Picklist ID found for attribute code: {attribute_code}. Skipping to next sheet.")
                                picklist_id = None
                            else:
                                picklist_id = picklist_ids.iloc[0]
       
                            # Determine the valid values for this picklist
                            filtered_valid_combinations = {code for pid, code in valid_combinations if pid == picklist_id}

                            # print("# Delete rows #")

                            # Define a function to filter rows and collect details of disregarded entries
                            def delete_combinations(row, filtered_valid_combinations):
                                # Check if the combination of 'Picklist ID' and the value in the second column is in the valid combinations
                                if (row.iloc[1]) not in filtered_valid_combinations:
                                    # Collect filename, sheet name, and the second column's value for disregarded entries
                                    disregarded_entries.append({
                                        'filename': filename,  # Assumes 'filename' is defined where this function is called
                                        'sheet_name': sheet_name,  # Assumes 'sheet_name' is defined where this function is called
                                        'value': row.iloc[1]
                                    })
                                    return False
                                return True
                            
                            
                            ##############################
                            ## Delete
                            ##############################
                            filtered_df = df[df.apply(delete_combinations, axis=1, args=(filtered_valid_combinations,))]

                            ##############################
                            ## Add
                            ##############################
                            # # Check if the DataFrame has at least two columns
                            if len(df.columns) < 2:     
                                print(f"Not enough columns in {filename} - {sheet_name}.")
                                # exit()
                                continue
                            
                            # Check if the DataFrame is empty
                            if not filtered_df.empty:
                                current_combinations = set(filtered_df.iloc[:, 1])
                            else:
                                current_combinations = set()
                                # print(f"No data in {filename} - {sheet_name}.")

                            # Determine missing combinations
                            missing_combinations = filtered_valid_combinations - current_combinations

                            # Define the language codes and their corresponding columns in gs1_df_picklists
                            language_mapping = {
                                "en_US": "Values in English used for user interface ",
                                "nl_NL": "Values in Dutch used for user interface ",
                                "nl_BE": "Values in Dutch used for user interface ",
                                "fr_FR": "Values in French used for user interface ",
                                "fr_BE": "Values in French used for user interface "
                            }
                    
                            for code_value in missing_combinations: 
                                added_entries.append({
                                        'filename': filename,  # Assumes 'filename' is defined where this function is called
                                        'sheet_name': sheet_name,  # Assumes 'sheet_name' is defined where this function is called
                                        'value': code_value
                                    })
                                
                                # Find the row in gs1_df_picklists corresponding to the current code_value and picklist_id
                                row_data = gs1_df_picklists[(gs1_df_picklists['Picklist ID'] == picklist_id) & (gs1_df_picklists['Code value'] == code_value)]

                                # Initialize the new row with the code value
                                new_row = {filtered_df.columns[1]: code_value}  # filtered_df.columns[1] should be the column where code values are stored

                                # Assign language specific values
                                for language_code, gs1_col in language_mapping.items():
                                    # Find columns in filtered_df that contain the language code
                                    matched_columns = [col for col in filtered_df.columns if language_code in col]
                                    # Get the value from gs1_df_picklists
                                    language_value = row_data[gs1_col].values[0] if not row_data.empty else ''
                                    # Assign this value to all matched columns in new_row
                                    for matched_col in matched_columns:
                                        new_row[matched_col] = language_value

                                # Append the new row to filtered_df
                                filtered_df = pd.concat([filtered_df, pd.DataFrame([new_row])], ignore_index=True)

                            # Make a total dataframe
                            if disregarded_entries or added_entries:
                                all_data_frames.append({'df': filtered_df, 'filename': filename, 'sheet_name': sheet_name, 'Picklist': picklist_id})

    return all_data_frames          
        
directory_path = 'Workfiles/LookupData'
final_lookupdata_df = read_all_excel_files(directory_path)
delete_lookupdata_df = final_lookupdata_df.copy()

####################
## Pre-calculations for possible additions
####################
print(f'### Pre-calculations for possible additions ###')

# Get relevant rows from attribute overview
gs1_df_attributes_processed = gs1_df_attributes.copy()

# Data Type and display type
def determine_types(row):
    format = row['Format']
    decimals = row['Deci-\nmals']
    if format == "Number":
        data_type = "Integer" if decimals == '0' else "Decimal"
        display_type = "NumericTextBox"
    elif format == "DateTime":
        data_type = "DateTime"
        display_type = "DateTime"
    elif format == "Text":
        data_type = "String"
        display_type = "TextBox" 
    elif format == "Picklist (T/F)":
        data_type = "String"
        display_type = "LookupTable"
    elif format == "Picklist":
        data_type = "String"
        display_type = "LookupTable"
    elif format == "NumberPicklist":
        data_type = "Integer" if decimals == '0' else "Decimal"
        display_type = "NumericTextBox"
    elif format == "Boolean":
        data_type = "String"
        display_type = "LookupTable"
    else:
        data_type = "Unknown"
        display_type = "Unknown"
    return pd.Series([data_type, display_type], index=['INPUT_Data_type', 'INPUT_Display_type'])

# Apply the function to each row of the dataframe
gs1_df_attributes_processed[['INPUT_Data_type', 'INPUT_Display_type']] = gs1_df_attributes_processed.apply(determine_types, axis=1)

# 
gs1_df_attributes_processed['INPUT_Attribute_name'] = gs1_df_attributes_processed['Attributename English'].apply(
                                                lambda x: x[:x.rfind('(')].strip() if '(' in x and x.endswith(')') else x.strip()
                                            ).apply(lambda x: f"CatSpec_{x}")


def clean_sheet_name(sheet_name):
    # Define invalid characters for Excel sheet names
    invalid_chars = ['\\', '/', '*', '[', ']', ':', '?', ' ']
    # Remove invalid characters
    for char in invalid_chars:
        sheet_name = sheet_name.replace(char, '')
    # Truncate to 30 characters
    return sheet_name

# Apply the function to the INPUT_Attribute_name column
gs1_df_attributes_processed['INPUT_Attribute_name'] = gs1_df_attributes_processed['INPUT_Attribute_name'].apply(clean_sheet_name)

# INPUT_Lookup_table_name
gs1_df_attributes_processed['INPUT_Lookup_table_name'] = np.select(
    [
        gs1_df_attributes_processed['Format'].isin(["Picklist (T/F)", "Boolean"]),
        gs1_df_attributes_processed['Format'] == "Picklist"
    ],
    [
        "YesNo",
        gs1_df_attributes_processed['INPUT_Attribute_name'].str.replace(r'\s+', '', regex=True).str.strip().apply(lambda x: x[8:][:30]) # max is 30 but need space for OneWs later on
    ],
    default=""
)

# Add INPUT_Allowed_uoms
code_value_concat = gs1_df_picklists.groupby('Picklist ID')['Code value'].apply(lambda x: '||'.join(x.dropna().astype(str))).rename('INPUT_Allowed_uoms')

# Using a left join ensures all original rows in gs1_df_attributes_processed are retained
gs1_df_attributes_processed = gs1_df_attributes_processed.merge(code_value_concat, on='Picklist ID', how='left')

# Fill NaNs with empty strings if any picklist IDs didn't have code values
gs1_df_attributes_processed['INPUT_Allowed_uoms'] = gs1_df_attributes_processed['INPUT_Allowed_uoms'].fillna('')


#Fill the table
gs1_df_attributes_processed['ID'] = ''
gs1_df_attributes_processed['Action'] = ''
gs1_df_attributes_processed['Unique Identifier'] = ''
gs1_df_attributes_processed['Attribute Type'] = 'Category'
gs1_df_attributes_processed['Attribute Name'] = gs1_df_attributes_processed['INPUT_Attribute_name']
gs1_df_attributes_processed['Attribute Long Name'] = gs1_df_attributes_processed['Attributename English']
gs1_df_attributes_processed['Attribute Parent Name'] = 'Category Specific Attributes'
gs1_df_attributes_processed['Data Type'] = gs1_df_attributes_processed['INPUT_Data_type']
gs1_df_attributes_processed['Display Type'] = gs1_df_attributes_processed['INPUT_Display_type']
gs1_df_attributes_processed['Is Collection'] = np.where(
                                            gs1_df_attributes_processed['Repeat'].str.len() > 0, 'YES', 'NO'
                                        )
gs1_df_attributes_processed['Is Inheritable'] = 'NO' #For OneWS is 'YES' bij picklisten en numberpicklisten
gs1_df_attributes_processed['Is Localizable'] = 'NO'
gs1_df_attributes_processed['Is Complex'] = 'NO'
gs1_df_attributes_processed['Is Lookup'] =  np.where(gs1_df_attributes_processed['INPUT_Display_type'] == 'LookupTable', 'YES', 'NO')
gs1_df_attributes_processed['Is Required'] = 'NO'
gs1_df_attributes_processed['Is ReadOnly'] = 'NO'
gs1_df_attributes_processed['Is Hidden'] = 'NO'
gs1_df_attributes_processed['Show At Entity Creation?'] = 'YES'
gs1_df_attributes_processed['Is Searchable'] = 'YES'
gs1_df_attributes_processed['Is Null Value Search Required'] = 'YES'
gs1_df_attributes_processed['Generate Report Table Column?'] = ''
gs1_df_attributes_processed['Default Value'] = ''
gs1_df_attributes_processed['Minimum Length'] = 0  
gs1_df_attributes_processed['Maximum Length'] = 0  
gs1_df_attributes_processed['Range From'] = ''
gs1_df_attributes_processed['Is Range From Inclusive'] = ''
gs1_df_attributes_processed['Range To'] = ''
gs1_df_attributes_processed['Is Range To Inclusive'] = ''
gs1_df_attributes_processed['Precision'] = gs1_df_attributes_processed['Precision'] = gs1_df_attributes_processed['Deci-\nmals'].replace('0', '')
gs1_df_attributes_processed['Use Arbitrary Precision?'] = np.where(
                                                        gs1_df_attributes_processed['Precision'].str.len() > 0, 'NO', ''
                                                    )
gs1_df_attributes_processed['UOM Type'] = np.where(gs1_df_attributes_processed['Format'] == 'NumberPicklist', 'Custom UOM', '') #numberbicklist ? --> "CustomUOM",  bij onews "gdsn uom"
gs1_df_attributes_processed['Allowed UOMs'] = np.where(gs1_df_attributes_processed['Format'] == 'NumberPicklist', gs1_df_attributes_processed['INPUT_Allowed_uoms'],'') #ONLY FOR numberbicklist
gs1_df_attributes_processed['Default UOM'] = np.where(gs1_df_attributes_processed['Format'] == 'NumberPicklist', gs1_df_attributes_processed['UoM fixed'],'')
gs1_df_attributes_processed['Allowable Values'] = ''
gs1_df_attributes_processed['LookUp Table Name'] = gs1_df_attributes_processed['INPUT_Lookup_table_name']
gs1_df_attributes_processed['Lookup Display Columns'] = gs1_df_attributes_processed['LookUp Table Name'].apply(lambda x: f"[{x}]" if x.strip() else x)
gs1_df_attributes_processed['Lookup Search Columns'] = gs1_df_attributes_processed['LookUp Table Name'].apply(lambda x: f"[{x}]" if x.strip() else x)
gs1_df_attributes_processed['Lookup Display Format'] = gs1_df_attributes_processed['LookUp Table Name'].apply(lambda x: f"[{x}]" if x.strip() else x)
gs1_df_attributes_processed['Lookup Sort Order'] = gs1_df_attributes_processed['LookUp Table Name'].apply(lambda x: f"[{x}]" if x.strip() else x)
gs1_df_attributes_processed['Lookup Order'] = gs1_df_attributes_processed['LookUp Table Name'].apply(lambda x: f"[{x}]" if x.strip() else x)
gs1_df_attributes_processed['Sort Order'] = 0
gs1_df_attributes_processed['Definition'] = ("GS1 Field_ID " + 
                                         gs1_df_attributes_processed['FieldID'].astype(str) + " " + 
                                         gs1_df_attributes_processed['Definition English'])
gs1_df_attributes_processed['Example'] = ''
gs1_df_attributes_processed['Business Rule'] = ''
gs1_df_attributes_processed['Label'] = ''
gs1_df_attributes_processed['Extension'] = ''
gs1_df_attributes_processed['Web URI'] = ''
gs1_df_attributes_processed['Enable History'] = 'YES'
gs1_df_attributes_processed['Apply Time Zone Conversion'] = 'NO'
gs1_df_attributes_processed['Attribute Regular Expression'] = ''
gs1_df_attributes_processed['Is UOM Localizable'] = 'NO'

gs1_df_attributes_processed.fillna('', inplace=True)

####################
## S7 
####################
print(f'### S7 ###')

# Select Category to baee changes on
maxeda_s7_df_category = maxeda_s7_df_scope[maxeda_s7_df_scope['Attribute Type'] == 'Category'].copy()
# Create a set from this
maxeda_attribute_s7_GS1ID_set = set(maxeda_s7_df_category['Attribute code'].replace('', np.nan).dropna())

    ####################
    ## Establish set of attributes for 1) additions, 2) deletions, and 3) overlapping 
    ####################

attribute_add_s7_GS1ID_set = gs1_attributes_GS1ID_set - maxeda_attribute_s7_GS1ID_set
attribute_delete_s7_GS1ID_set = maxeda_attribute_s7_GS1ID_set - gs1_attributes_GS1ID_set
attribute_overlap_s7_GS1ID_set = gs1_attributes_GS1ID_set & maxeda_attribute_s7_GS1ID_set

    ####################
    ## Delete
    ####################

def delete_attributes(delete_set, df, return_df):

    temp_df = df[df['Attribute code'].isin(delete_set)].copy()
    temp_df['Action'] = 'Delete'
    return_df = pd.concat([return_df, temp_df], ignore_index=True)
    return return_df

print(f'## Delete ##')

# Create an empty delete_attributes_s7_df with necessary columns
delete_attributes_s7_df = pd.DataFrame(columns=list(maxeda_s7_df_scope.columns))

delete_attributes_s7_df = delete_attributes(attribute_delete_s7_GS1ID_set, maxeda_s7_df_scope, delete_attributes_s7_df)
# print(delete_attributes_s7_df)

attribute_delete_s7_MaxedaIDs_set = set(delete_attributes_s7_df['ID'].dropna())
attribute_delete_s7_LookupTableName_set = set(delete_attributes_s7_df['LookUp Table Name'].dropna())
if 'YesNo' in attribute_delete_s7_LookupTableName_set:
    attribute_delete_s7_LookupTableName_set.remove('YesNo')

    ####################
    ## Additions
    ####################
print(f'## Add ##')
def add_attributes_s7(add_set, all_additions_attributes_s7_df):

    additions_attributes_s7_df = gs1_df_attributes_processed[gs1_df_attributes_processed['FieldID'].isin(add_set)].copy()

    # OneWs attributes
    additions_attributes_onews_s7_df = additions_attributes_s7_df.copy()

    # Setting 'Attribute Type' to 'Common' for all rows
    additions_attributes_onews_s7_df['Attribute Type'] = 'Common'

    # Replacing 'CatSpec' with 'OneWS' in 'Attribute Name' and removing all spaces
    additions_attributes_onews_s7_df['Attribute Name'] = additions_attributes_onews_s7_df['Attribute Name'].str.replace('CatSpec', 'OneWS').str.replace(' ', '')

    # Adjust Is Inheritable for OneWs
    additions_attributes_onews_s7_df['Is Inheritable'] = np.where(
        (additions_attributes_s7_df['Format'] == 'NumberPicklist') | (additions_attributes_s7_df['Format'] == 'Picklist'),
        'YES',  # Set to 'YES' if 'Format' is either 'NumberPicklist' or 'Picklist'
        additions_attributes_onews_s7_df['Is Inheritable']  # Keep the original value if conditions are False
    )

    # Update 'Data Type' and 'Display Type' based on 'Format'
    additions_attributes_s7_df['Data Type'] = np.where(
        additions_attributes_s7_df['Format'] == 'Boolean',
        'Boolean',  # Set 'Data Type' to 'Boolean' if 'Format' is 'Boolean'
        additions_attributes_s7_df['Data Type']  # Otherwise, keep the current value
    )

    additions_attributes_s7_df['Display Type'] = np.where(
        additions_attributes_s7_df['Format'] == 'Boolean',
        'DropDown',  # Set 'Display Type' to 'DropDown' if 'Format' is 'Boolean'
        additions_attributes_s7_df['Display Type']  # Otherwise, keep the current value
    )

    # Adjust Is Inheritable 
    additions_attributes_onews_s7_df['Is Inheritable'] = np.where(
        (additions_attributes_s7_df['Format'] == 'NumberPicklist') | (additions_attributes_s7_df['Format'] == 'Picklist'),
        'YES',  # Set to 'YES' if 'Format' is either 'NumberPicklist' or 'Picklist'
        additions_attributes_onews_s7_df['Is Inheritable']  # Keep the original value if conditions are False
    )

    additions_attributes_onews_s7_df['LookUp Table Name'] = additions_attributes_onews_s7_df['Attribute Name'].str.replace('_', '').str.slice(0, 30)
    additions_attributes_onews_s7_df['LookUp Display Columns'] = '[Code],[Description]'
    additions_attributes_onews_s7_df['LookUp Search Columns'] = '[Code],[Description]'
    additions_attributes_onews_s7_df['LookUp Display Format'] = '[Code]'
    additions_attributes_onews_s7_df['LookUp Sort Order'] = '[Code]'
    additions_attributes_onews_s7_df['Export Format'] = '[Code]'

    all_additions_attributes_s7_df = pd.concat([all_additions_attributes_s7_df, additions_attributes_s7_df, additions_attributes_onews_s7_df], ignore_index=True)
    
    return all_additions_attributes_s7_df

# Create an empty delete_attributes_s7_df with necessary columns
all_additions_attributes_s7_df = pd.DataFrame(columns=list(gs1_df_attributes_processed.columns))

all_additions_attributes_s7_df = add_attributes_s7(attribute_add_s7_GS1ID_set, all_additions_attributes_s7_df)

s7_add_picklist_ID_set = set(zip(
    all_additions_attributes_s7_df[
        (all_additions_attributes_s7_df['Picklist ID'].notna()) & 
        (all_additions_attributes_s7_df['Picklist ID'] != '') & 
        (all_additions_attributes_s7_df['LookUp Table Name'].notna()) & 
        (all_additions_attributes_s7_df['LookUp Table Name'] != '') & 
        (all_additions_attributes_s7_df['LookUp Table Name'] != 'YesNo') & 
        (all_additions_attributes_s7_df['Display Type'] == 'LookupTable')
    ]['Picklist ID'],
    all_additions_attributes_s7_df[
        (all_additions_attributes_s7_df['Picklist ID'].notna()) & 
        (all_additions_attributes_s7_df['Picklist ID'] != '') & 
        (all_additions_attributes_s7_df['LookUp Table Name'].notna()) & 
        (all_additions_attributes_s7_df['LookUp Table Name'] != '') & 
        (all_additions_attributes_s7_df['LookUp Table Name'] != 'YesNo') &
        (all_additions_attributes_s7_df['Display Type'] == 'LookupTable')
    ]['LookUp Table Name']
))
# print(f"s7_add_picklist_ID_set: {s7_add_picklist_ID_set}")

#########
## Questions
#########

    ####################
    ## Changes
    ####################
print(f'## Changes ##')
overlap_attributes_df = gs1_df_attributes_processed[gs1_df_attributes_processed['FieldID'].isin(attribute_overlap_s7_GS1ID_set)].copy()

# Step 1: Duplicate and prefix selected columns in maxeda_s7_df to compare including the key being attribute code
columns_to_compare = ['Attribute code','Data Type', 'Display Type', 'Precision', 'Allowed UOMs', 'Default UOM', 'Is Collection', 'LookUp Table Name']

# Create a new DataFrame, maxeda_s7_changes_df, with only the specified columns from maxeda_s7_df
maxeda_s7_changes_df = maxeda_s7_df_category[columns_to_compare].copy()

# Remove 'Attribute code' to function as the key to join on later
columns_to_compare.remove('Attribute code')

# Prefix check-columns
rename_dict = {col: 'ORIGINAL_' + col for col in columns_to_compare}

# Rename the columns using the dictionary
maxeda_s7_changes_df.rename(columns=rename_dict, inplace=True)

# The original lookup table is needed for determination later on not to compare
columns_to_compare.remove('LookUp Table Name')

# Step 2: Perform a left join from overlap_attributes_df to the modified maxeda_s7_df
merged_df = overlap_attributes_df.merge(
    maxeda_s7_changes_df,
    left_on='FieldID',
    right_on='Attribute code',
    how='left'
)

# Step 3: Compare original and newly added values, collect discrepancies and reasons
discrepancy_details = []
for col in columns_to_compare:
    for index, row in merged_df.iterrows():
        # Trim values and check if they are not empty
        original_trimmed = (str(row['ORIGINAL_' + col]).strip() if pd.notnull(row['ORIGINAL_' + col]) else '')
        new_trimmed = (str(row[col]).strip() if pd.notnull(row[col]) else '')
        
        # Add to discrepancy details only if both are not empty and different
        if (original_trimmed or new_trimmed) and (original_trimmed != new_trimmed):
            discrepancy_details.append({
                'FieldID': row['FieldID'],
                'Column': col,
                'Original Value': row['ORIGINAL_' + col],
                'New Value': row[col],
                'Picklist ID': row['Picklist ID']
            })


# Convert discrepancy details to a DataFrame for better visualization and analysis
discrepancy_df = pd.DataFrame(discrepancy_details)

# Step 4: Extract unique FieldIDs with discrepancies
change_set = set(discrepancy_df['FieldID'].dropna())

# Filter the dataset on the the changed items
changes_s7_df = merged_df[merged_df['FieldID'].isin(change_set)].copy()

# Make set for picklist-value additions
s7_change_to_picklist_ID_set = set(zip(
    changes_s7_df[
        (changes_s7_df['Picklist ID'].notna()) & 
        (changes_s7_df['Picklist ID'] != '') & 
        (changes_s7_df['LookUp Table Name'].notna()) & 
        (changes_s7_df['LookUp Table Name'] != '') & 
        (changes_s7_df['LookUp Table Name'] != 'YesNo') &
        (changes_s7_df['ORIGINAL_Display Type'] != 'LookupTable') & 
        (changes_s7_df['Display Type'] == 'LookupTable')
    ]['Picklist ID'],
    changes_s7_df[
        (changes_s7_df['Picklist ID'].notna()) & 
        (changes_s7_df['Picklist ID'] != '') & 
        (changes_s7_df['LookUp Table Name'].notna()) & 
        (changes_s7_df['LookUp Table Name'] != '') & 
        (changes_s7_df['ORIGINAL_Display Type'] != 'LookupTable') & 
        (changes_s7_df['Display Type'] == 'LookupTable')
    ]['LookUp Table Name']
))

s7_change_from_LookupTableName_set = set(
    changes_s7_df[
        (changes_s7_df['ORIGINAL_LookUp Table Name'] != 'YesNo') & 
        (changes_s7_df['ORIGINAL_Display Type'] == 'LookupTable') & 
        (changes_s7_df['Display Type'] != 'LookupTable')
    ]['ORIGINAL_LookUp Table Name']
)

# print(f"s7_change_to_picklist_set: {s7_change_to_picklist_ID_set}")

# Combine the two DataFrames into one new DataFrame
final_s7_df = pd.concat([delete_attributes_s7_df, all_additions_attributes_s7_df, changes_s7_df], ignore_index=True)
final_s7_additions_changes_df = pd.concat([all_additions_attributes_s7_df, changes_s7_df], ignore_index=True)

# Filter columns for the output
columns_s7 = [
    "ID", "Action", "Unique Identifier", "Attribute Type", "Attribute Name",
    "Attribute Long Name", "Attribute Parent Name", "Data Type", "Display Type",
    "Is Collection", "Is Inheritable", "Is Localizable", "Is Complex", "Is Lookup",
    "Is Required", "Is ReadOnly", "Is Hidden", "Show At Entity Creation?", "Is Searchable",
    "Is Null Value Search Required", "Generate Report Table Column?", "Default Value",
    "Minimum Length", "Maximum Length", "Range From", "Is Range From Inclusive",
    "Range To", "Is Range To Inclusive", "Precision", "Use Arbitrary Precision?",
    "UOM Type", "Allowed UOMs", "Default UOM", "Allowable Values", "LookUp Table Name",
    "Lookup Display Columns", "Lookup Search Columns", "Lookup Display Format",
    "Lookup Sort Order", "Export Format", "Sort Order", "Definition", "Example",
    "Business Rule", "Label", "Extension", "Web URI", "Enable History",
    "Apply Time Zone Conversion", "Attribute Regular Expression", "Is UOM Localizable"
]

final_s7_df = final_s7_df[columns_s7]
delete_attributes_s7_df = delete_attributes_s7_df[columns_s7]
final_s7_additions_changes_df = final_s7_additions_changes_df[columns_s7]

####################
## S8 
####################

print(f'### S8  ###')

    ####################
    ## Delete
    ####################
print(f'## Delete ##')

# Initiate the dataset
delete_attributes_s8_df = maxeda_s8_df_scope[maxeda_s8_df_scope['ID'].isin(attribute_delete_s7_MaxedaIDs_set)].copy()
delete_attributes_s8_df['Action'] = 'Delete'

    ####################
    ## Additions
    ####################
print(f'## Add ##')

# Initiate the dataset
additions_attributes_s8_df = gs1_df_attributes_processed[gs1_df_attributes_processed['FieldID'].isin(attribute_add_s7_GS1ID_set)].copy()

# Contruct 'Attribute Path'
additions_attributes_s8_df['Attribute Path'] = additions_attributes_s8_df['Attribute Parent Name'] + '//' + additions_attributes_s8_df['Attribute Name']

# Create an empty delete_attributes_s7_df with necessary columns
additons_s8_locale_df = pd.DataFrame(columns=list(additions_attributes_s8_df.columns))

# Create a DataFrame with repeated rows, each original row appearing four times for different locales
locales = ['nl_BE', 'nl_NL', 'fr_BE', 'fr_FR']
additons_s8_locale_df = pd.DataFrame(
    np.repeat(additions_attributes_s8_df.values, len(locales), axis=0),  # Repeat each row
    columns=additions_attributes_s8_df.columns
)

additons_s8_locale_df['Locale'] = np.tile(locales, len(additions_attributes_s8_df))  # Assign locales repeatedly for each group of rows

# Define 'Attribute Long Name' based on 'Locale'
additons_s8_locale_df['Attribute Long Name'] = np.where(
    additons_s8_locale_df['Locale'].isin(['nl_BE', 'nl_NL']),
    additions_attributes_s8_df['Attributename Dutch'].repeat(len(locales)).values,  # Repeat values to match the new row expansion
    additions_attributes_s8_df['Attributename French'].repeat(len(locales)).values
)

#Rename headers
additons_s8_locale_df.rename(columns={'Minimum Length': 'Min Length', 'Maximum Length': 'Max Length'}, inplace=True)

# Empty all values in 'Min Length' and 'Max Length'
additons_s8_locale_df['Min Length'] = ''
additons_s8_locale_df['Max Length'] = ''

# Onews
additions_onews_s8_locale_df = additons_s8_locale_df.copy()
# Replacing 'CatSpec' with 'OneWS' in 'Attribute Name' and removing all spaces
additions_onews_s8_locale_df['Attribute Name'] = additions_onews_s8_locale_df['Attribute Name'].str.replace('CatSpec', 'OneWS').str.replace(' ', '').str.replace('Category Specific Attributes//', 'OneWS_XXXX//')

all_additions_attributes_s8_df = pd.concat([additons_s8_locale_df, additions_onews_s8_locale_df], ignore_index=True)

# Find the IDs that are in attribute_delete_s7_MaxedaIDs_set but not in maxeda_s8_IDs_set
missing_ids = attribute_delete_s7_MaxedaIDs_set - maxeda_s8_IDs_set

# # Print the missing IDs
# print("IDs in attribute_delete_s7_MaxedaIDs_set that are not in maxeda_s8_df:")
# for id in missing_ids:
#     print(id) 


# Combine all data frames
final_s8_df = pd.concat([delete_attributes_s8_df, all_additions_attributes_s8_df], ignore_index=True)

# Configure output
columns_s8 = ['ID', 'Action', 'Unique Identifier', 'Attribute Path', 'Locale', 'Attribute Long Name', 'Min Length', 'Max Length', 'Definition', 'Example', 'Business Rule']
final_s8_df = final_s8_df[columns_s8]
delete_attributes_s8_df = delete_attributes_s8_df[columns_s8]
all_additions_attributes_s8_df = all_additions_attributes_s8_df[columns_s8]

#######################
## S23 
#######################

# Combine the sets of deleted picklists and exiting attributes changed format FROM Picklist to something else
LookupTable_delete_total_set = attribute_delete_s7_LookupTableName_set.union(s7_change_from_LookupTableName_set)
# Combine the sets of added picklists and exiting attributes changed format TO Picklist from something else
LookupTable_add_total_set = s7_add_picklist_ID_set.union(s7_change_to_picklist_ID_set)

# Create a new DataFrame with the same columns as maxeda_s23_df
maxeda_s23_total_df = pd.DataFrame(columns=maxeda_s23_df.columns)

    #######################
    ## Delete
    #######################

maxeda_s23_delete_df = maxeda_s23_df[maxeda_s23_df['Table Name'].isin(LookupTable_add_total_set)].copy()
maxeda_s23_delete_df['Action'] = 'Delete'

maxeda_s23_total_df = pd.concat([maxeda_s23_total_df, maxeda_s23_delete_df], ignore_index=True)


    #######################
    ## Add
    #######################

for picklist_id, lookup_table_name in LookupTable_add_total_set:
    
    # Determine the common row values
    table_name = lookup_table_name
    sequence = 0
    column_name_first_row = "Code"
    column_name_second_row = lookup_table_name

    # Find the corresponding data type and precision from gs1_df_attributes_processed
    match = gs1_df_attributes_processed[gs1_df_attributes_processed['Picklist ID'] == picklist_id]
    data_type = match['Data Type'].values[0] if not match.empty else None
    
    width = 500
    precision = 0
    nullable_first_row = "Yes"
    nullable_second_row = "Yes"
    is_unique_first_row = "Yes"
    is_unique_second_row = "No"
    
    if lookup_table_name.startswith("OneWS"):
        column_name_second_row = "Description"
        width = 250
        nullable_first_row = "No"

    # Create the first row
    first_row = {
        'Table Name': table_name,
        'Sequence': sequence,
        'Column Name': column_name_first_row,
        'Data Type': data_type,
        'Width': width,
        'Precision': precision,
        'Nullable?': nullable_first_row,
        'Is Unique': is_unique_first_row
    }
    
    # Create the second row
    second_row = {
        'Table Name': table_name,
        'Sequence': sequence,
        'Column Name': column_name_second_row,
        'Data Type': data_type,
        'Width': width,
        'Precision': precision,
        'Nullable?': "Yes",
        'Is Unique': is_unique_second_row
    }
    
    # Append the rows to the new DataFrame
    # maxeda_s23_total_df = maxeda_s23_total_df.append(first_row, ignore_index=True)
    # maxeda_s23_total_df = maxeda_s23_total_df.append(second_row, ignore_index=True)
    
    all_additions_attributes_s23_df = pd.DataFrame(columns=list(maxeda_s23_delete_df.columns))
    all_additions_attributes_s23_df = pd.concat([all_additions_attributes_s23_df, pd.DataFrame([first_row]), pd.DataFrame([second_row])], ignore_index=True)


    maxeda_s23_total_df = pd.concat([maxeda_s23_total_df, all_additions_attributes_s23_df], ignore_index=True)
 
    # print(maxeda_s23_total_df)


#######################
## New lookup attributes data for new picklists
#######################

# Loop over the set of tuples
for picklist_id, lookup_table_name in LookupTable_add_total_set:

    # Filter gs1_df_picklists for rows with the current Picklist ID
    new_picklist_values_df = gs1_df_picklists[gs1_df_picklists['Picklist ID'] == picklist_id].copy()

    # Create a new DataFrame with the required columns and headers
    reconfigured_df = pd.DataFrame({
        'Id': [''] * len(new_picklist_values_df),
        'Code': new_picklist_values_df['Code value'],
        f'{lookup_table_name}//en_US': new_picklist_values_df['Values in English used for user interface '],
        f'{lookup_table_name}//nl_BE': new_picklist_values_df['Values in Dutch used for user interface '],
        f'{lookup_table_name}//nl_NL': new_picklist_values_df['Values in Dutch used for user interface '],
        f'{lookup_table_name}//fr_BE': new_picklist_values_df['Values in French used for user interface '],
        f'{lookup_table_name}//fr_FR': new_picklist_values_df['Values in French used for user interface ']
    })

    # print(f"lookup table name: {lookup_table_name}")

    # Add to final dataframe 
    final_lookupdata_df.append({'df': reconfigured_df, 'filename': 'NO FILE: addition', 'sheet_name': lookup_table_name, 'Picklist': picklist_id})



###################
## Write output
###################


print('## Metadata ##')
# Create a DataFrame to store data on first metadata sheet

def metadata_lookupvalues(dataframe):
    metadata_sheet = {
        'LookupTableName': [item['sheet_name'] for item in dataframe],
        'SheetName': [item['sheet_name'] for item in dataframe],
        'Load Lookup?': ['Yes'] * len(dataframe)
    }
    metadata_sheet_df = pd.DataFrame(metadata_sheet)

    return metadata_sheet_df

def metadata_s7_s8_s23(physical_sheet_name):
    # Split physical_sheet_name into parts before and after the first "-"
    parts = physical_sheet_name.split("-", 1)
    sheet_no = parts[0].strip() if len(parts) > 0 else ""
    data_model_type_name = parts[1].strip() if len(parts) > 1 else ""

    # Create the metadata dictionary
    metadata_sheet = {
        'Sheet No': [sheet_no],
        'DataModel Type Name': [data_model_type_name],
        'Physical Sheet Name': [physical_sheet_name],
        'Load Lookup?': ['Yes']
    }

    # Convert the dictionary to a DataFrame
    metadata_sheet_df = pd.DataFrame(metadata_sheet)

    return metadata_sheet_df

output_file_path_attributes = os.path.join(os.getcwd(), 'GS1_vs_Datamodel_Comparison_Attributes.xlsx')
output_file_path_lookupdata = os.path.join(os.getcwd(), 'LookupData.xlsx')

############################
## Combined output
############################
print('## Output writer ##')
# Write S7, S8 and S23
with pd.ExcelWriter(output_file_path_attributes, engine='openpyxl') as writer:
    print("## S7 ##")
    final_s7_df.to_excel(writer, sheet_name='S7 - Attribute', index=False)
    print("## S8 ##")
    final_s8_df.to_excel(writer, sheet_name='S8 - Attribute - Locale', index=False)
    print("## S23 ##")
    maxeda_s23_total_df.to_excel(writer, sheet_name='S23 - Lookup Model', index=False)   


# Write lookup tabale values
with pd.ExcelWriter(output_file_path_lookupdata, engine='openpyxl') as writer:
    print("## LookupData ##")

    # Create metadata data frame
    metadata_lookupvalues_total = metadata_lookupvalues(final_lookupdata_df)
    # Write the metadata DataFrame as the first sheet named 'Metadata'
    metadata_lookupvalues_total.to_excel(writer, sheet_name='Metadata', index=False)

    # Write each DataFrame to its respective sheet
    for item in tqdm(final_lookupdata_df, desc="Writing lookupdata sheets"):
        # Write DataFrame to a sheet named after the original sheet_name
        item['df'].to_excel(writer, sheet_name=item['sheet_name'], index=False)

############################
## Output in workflow
############################

# Write deletions of lookup data values
with pd.ExcelWriter(os.path.join(os.getcwd(), '1_Delete_LookupData_Values.xlsx'), engine='openpyxl') as writer:
    print("## 1.1 - Deletions - LookupData delete values##")
 
    # Create metadata data frame
    metadata_lookupvalues_delete = metadata_lookupvalues(delete_lookupdata_df)   
    # Write the metadata DataFrame as the first sheet named 'Metadata'
    metadata_lookupvalues_delete.to_excel(writer, sheet_name='Metadata', index=False)

    # Write each DataFrame to its respective sheet
    for item in tqdm(delete_lookupdata_df, desc="Writing lookupdata sheets"):
        # Write DataFrame to a sheet named after the original sheet_name
        item['df'].to_excel(writer, sheet_name=item['sheet_name'], index=False)


with pd.ExcelWriter(os.path.join(os.getcwd(), '2_Delete_LookupData_Tables_S23.xlsx'), engine='openpyxl') as writer:
    print("## 1.2 - Deletions - LookUp Table S23 ##")
    
    # Create metadata data frame
    metadata_s23_delete = metadata_s7_s8_s23('S23 - Lookup Model')   
    # Write the metadata DataFrame as the first sheet named 'Metadata'
    metadata_s23_delete.to_excel(writer, sheet_name='Metadata', index=False)
    
    maxeda_s23_delete_df.to_excel(writer, sheet_name='S23 - Lookup Model', index=False)


with pd.ExcelWriter(os.path.join(os.getcwd(), '3_Delete_Attributes_S8.xlsx'), engine='openpyxl') as writer:
    print("## 1.3 - Deletions - Attributes S8 ##")

    # Create metadata data frame
    metadata_s8_delete = metadata_s7_s8_s23('S8 - Attribute - Locale')   
    # Write the metadata DataFrame as the first sheet named 'Metadata'
    metadata_s8_delete.to_excel(writer, sheet_name='Metadata', index=False)

    delete_attributes_s8_df.to_excel(writer, sheet_name='S8 - Attribute - Locale', index=False)

with pd.ExcelWriter(os.path.join(os.getcwd(), '4_Delete_Attributes_S7.xlsx'), engine='openpyxl') as writer:
    print("## 1.4 - Deletions - Attributes S7 ##")

    # Create metadata data frame
    metadata_s7_delete = metadata_s7_s8_s23('S7 - Attribute')   
    # Write the metadata DataFrame as the first sheet named 'Metadata'
    metadata_s7_delete.to_excel(writer, sheet_name='Metadata', index=False)

    delete_attributes_s7_df.to_excel(writer, sheet_name='S7 - Attribute', index=False)

############ Delete Bricks, Families, Segments & Add Bricks, Families, Segments 

with pd.ExcelWriter(os.path.join(os.getcwd(), '7_Add_Attributes_S7.xlsx'), engine='openpyxl') as writer:
    print("## 2.2 - Additions - Attributes S7 ##")

    # Create metadata data frame
    metadata_s7_add = metadata_s7_s8_s23('S7 - Attribute')   
    # Write the metadata DataFrame as the first sheet named 'Metadata'
    metadata_s7_add.to_excel(writer, sheet_name='Metadata', index=False)

    final_s7_additions_changes_df.to_excel(writer, sheet_name='S7 - Attribute', index=False)

with pd.ExcelWriter(os.path.join(os.getcwd(), '8_Add_Attributes_S8.xlsx'), engine='openpyxl') as writer:
    print("## 2.2 - Additions - Attributes S8 ##")

    # Create metadata data frame
    metadata_s8_add = metadata_s7_s8_s23('S8 - Attribute - Locale')   
    # Write the metadata DataFrame as the first sheet named 'Metadata'
    metadata_s8_add.to_excel(writer, sheet_name='Metadata', index=False)

    all_additions_attributes_s8_df.to_excel(writer, sheet_name='S8 - Attribute - Locale', index=False)

with pd.ExcelWriter(os.path.join(os.getcwd(), '9_Add_LookupData_Tables_S23.xlsx'), engine='openpyxl') as writer:
    print("## 2.3 - Additions - LookUp Table S23 ##")

    # Create metadata data frame
    metadata_s23_add = metadata_s7_s8_s23('S23 - Lookup Model')   
    # Write the metadata DataFrame as the first sheet named 'Metadata'
    metadata_s23_add.to_excel(writer, sheet_name='Metadata', index=False)

    all_additions_attributes_s23_df.to_excel(writer, sheet_name='S23 - Lookup Model', index=False)

with pd.ExcelWriter(os.path.join(os.getcwd(), '10_Add_LookupData_Values.xlsx'), engine='openpyxl') as writer:
    print("## 2.4 - Additions - LookupData values##")
    
    # Filter the list for items with 'filename' == 'NO FILE: addition'
    lookupvalues_add_df = [item for item in final_lookupdata_df if item['filename'] == 'NO FILE: addition']

    # Create metadata data frame
    metadata_lookupvalues_add = metadata_lookupvalues(lookupvalues_add_df)   
    # Write the metadata DataFrame as the first sheet named 'Metadata'
    metadata_lookupvalues_add.to_excel(writer, sheet_name='Metadata', index=False)

    # Write each DataFrame to its respective sheet if 'filename' is 'NO FILE: addition'
    for item in tqdm(lookupvalues_add_df, desc="Writing lookupdata sheets"):
        # Write DataFrame to a sheet named after the original sheet_name
        item['df'].to_excel(writer, sheet_name=item['sheet_name'], index=False)



####################
## Testing
####################            
# Load the updated Excel file into a DataFrame to confirm it saved correctly
loaded_attributes_s7_df = pd.read_excel(output_file_path_attributes, sheet_name='S7 - Attribute')
loaded_attributes_s8_df = pd.read_excel(output_file_path_attributes, sheet_name='S8 - Attribute - Locale')
loaded_lookupdata_df_metadata = pd.read_excel(output_file_path_lookupdata, sheet_name='Metadata')

Expected_s7_additions = 58
Expected_s7_deletions = 287
Expected_s7_changes = 710
loaded_attributes_s7 = 1191 # vs. calc >1113 (2* additions - Category & Common OneWS, 1.X * delete - OFTEN MISSING attribute code for Common OneWS, Changes - 1.X * change, some Attribute code have multiple rows)


Expected_s8_additions = Expected_s7_additions * 4 * 2 # languages + OneWS
Expected_attribute_delete_s7_MaxedaIDs_set = 329 # vs calc 287 * (1 + X). ? Because only 2 Common OneWS have attribute code incorporated, some ID's have multiple rows
Expected_s8_deletions = Expected_attribute_delete_s7_MaxedaIDs_set * 4 # languages
loaded_attributes_s8 = Expected_s8_additions + Expected_s8_deletions # vs. calc 1780  (Expected_s8_additions + (4 * delete considering direct ID'))

expected_items_metadata_lookupdata = 373
loaded_sheets_lookupdata = 374

assert len(attribute_add_s7_GS1ID_set) == Expected_s7_additions, f"Expected {Expected_s7_additions} additions, got {len(attribute_add_s7_GS1ID_set)}"
assert len(attribute_delete_s7_GS1ID_set) == Expected_s7_deletions, f"Expected {Expected_s7_deletions} deletions, got {len(attribute_delete_s7_GS1ID_set)}"
assert len(change_set) == Expected_s7_changes, f"Expected {Expected_s7_changes} changes, got {len(change_set)}"
assert len(loaded_attributes_s7_df) == loaded_attributes_s7, f"Expected {loaded_attributes_s7} total entries in S7, got {len(loaded_attributes_s7_df)}"

assert len(all_additions_attributes_s8_df) == Expected_s8_additions, f"Expected {Expected_s8_additions} additions locale in S8, got {len(all_additions_attributes_s8_df)}"
assert len(attribute_delete_s7_MaxedaIDs_set) == Expected_attribute_delete_s7_MaxedaIDs_set , f"Expected {Expected_attribute_delete_s7_MaxedaIDs_set} in ID-set, got {len(attribute_delete_s7_MaxedaIDs_set)}"
assert len(loaded_attributes_s8_df) == loaded_attributes_s8, f"Expected {loaded_attributes_s8} total entries in S8, got {len(loaded_attributes_s8_df)}"

assert len(loaded_lookupdata_df_metadata) == expected_items_metadata_lookupdata, f"Metadata sheet should contain exactly {expected_items_metadata_lookupdata} items, found {len(loaded_lookupdata_df_metadata)} items."
actual_sheets_count = len(pd.ExcelFile(output_file_path_lookupdata).sheet_names)
assert actual_sheets_count == loaded_sheets_lookupdata, f"Excel file should contain {loaded_sheets_lookupdata} sheets, found {actual_sheets_count} sheets."
# Check if the third sheet is named 'TOclothingforpets'
third_sheet_name = pd.ExcelFile(output_file_path_lookupdata).sheet_names[2]  # Indexing starts from 0
assert third_sheet_name == 'TOclothingforpets', f"The third sheet is named {third_sheet_name}, not 'TOclothingforpets'."
# Load the third sheet
third_sheet_df = pd.read_excel(output_file_path_lookupdata, sheet_name=third_sheet_name)
# Assert that this sheet has 28 items (rows)
assert len(third_sheet_df) == 28, f"The sheet 'TOclothingforpets' has {len(third_sheet_df)} items, not 28."


print("All tests passed successfully.")
