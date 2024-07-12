import pandas as pd
import os
from dotenv import load_dotenv
import numpy as np
import re
import openpyxl
from tqdm import tqdm  # Import tqdm
import time
import psutil

# Function to measure CPU and memory usage
def get_cpu_memory_usage():
    process = psutil.Process(os.getpid())
    memory_info = process.memory_info()
    cpu_usage = psutil.cpu_percent(interval=None)
    memory_usage = memory_info.rss / (1024 ** 2)  # Convert bytes to MB
    return cpu_usage, memory_usage

# Start timing
start_time = time.time()

# Start measuring CPU and memory before the code block
start_cpu_usage, start_memory_usage = get_cpu_memory_usage()

load_dotenv()

# Function to extract individual parent categories from the Parent Category Path
def extract_parents(parent_path):
    # Remove the "GS1//" prefix if it exists
    if parent_path.startswith("GS1//"):
        parent_path = parent_path[len("GS1//"):]
    # Split by "//" to get individual parent categories
    parts = parent_path.split('//')
    # Extract Segment, Family, and Class codes with safe indexing
    return (
        parts[0] if len(parts) > 0 else '',
        parts[1] if len(parts) > 1 else '',
        parts[2] if len(parts) > 2 else ''
    )

def create_category_df(categories, level, sheet):
    # print(f'## Create categories at {level} level ##')

    # Initialize an empty DataFrame for categories with specified columns
    columns = sheet.drop(columns=['Segment Code', 'Family Code', 'Class Code']).columns
    Category = pd.DataFrame(columns=columns)
    name_col = level.capitalize() + ' Code'
    path_cols = get_path_cols(level)
    
    # Assign sorted codes to the Category Name column
    Category['Category Name'] = sorted(categories)

    # Define locale settings based on the sheet
    if sheet.equals(maxeda_s9):
        locales = [{'title_prefix': '', 'locale': ''}]
    elif sheet.equals(maxeda_s10):
        locales = [
            {'title_prefix': 'NL ', 'locale': 'nl_NL'},
            {'title_prefix': 'FR ', 'locale': 'fr_FR'}
        ]

    # Process each locale setting
    all_categories = pd.DataFrame()
    for setting in locales:
        _Category = process_locale_setting(Category.copy(), setting, name_col, path_cols, level)
        all_categories = pd.concat([all_categories, _Category], ignore_index=True)
    
    if sheet.equals(maxeda_s10):
        # Duplicate for 'BE' locales
        be_categories = all_categories.copy()
        be_categories['Locale'] = be_categories['Locale'].replace({'nl_NL': 'nl_BE', 'fr_FR': 'fr_BE'})
        all_categories = pd.concat([all_categories, be_categories], ignore_index=True)

    return all_categories

def get_path_cols(level):
    """Define path columns based on category level."""
    if level == 'brick':
        return ['Segment Code', 'Family Code', 'Class Code']
    elif level == 'family':
        return ['Segment Code']
    elif level == 'class':
        return ['Segment Code', 'Family Code']
    return []

def process_locale_setting(Category, setting, name_col, path_cols, level):
    """Process a single locale setting and return the updated DataFrame."""
    title_col = setting['title_prefix'] + level.capitalize() + ' Title'
    if setting['locale']:
        Category['Locale'] = setting['locale']
    
    Category['Category Long Name'] = Category['Category Name'].map(
        lambda x: f"{x} - {gs1_df.loc[gs1_df[name_col].astype(str) == x, title_col].iloc[0]}"
        if x in gs1_df[name_col].astype(str).values else ''
    )
    
    Category['Parent Category Path'] = Category['Category Name'].apply(
        lambda x: f"GS1//{'//'.join(gs1_df.loc[gs1_df[name_col].astype(str) == x, path_cols].astype(str).iloc[0])}"
        if x in gs1_df[name_col].astype(str).values else 'GS1'
    )

    Category['Hierarchy Name'] = 'GS1 Hierarchy'

    # Duplicate for 'GPC' hierarchy
    Category_dup = Category.copy()
    Category_dup['Category Long Name'] = Category_dup['Category Long Name'].str.split(' - ').str[1]
    Category_dup['Parent Category Path'] = Category_dup['Parent Category Path'].str.replace('GS1//', '').str.replace('GS1', '')
    Category_dup['Hierarchy Name'] = 'GPC'

    return pd.concat([Category, Category_dup], ignore_index=True)

def delete_category_df(categories, sheet):
    # print(f'### Delete {sheet} ###')    
    # Filter the datamodel on the rows that need deletions
    Delete_categories = sheet[sheet['Category Name'].isin(categories)].copy()
    
    # Substitute nan by blank
    Delete_categories.replace('nan', '', inplace=True)

    # Add the delete action
    Delete_categories['Action'] = 'Delete'
    
    return Delete_categories

# Set file paths and read the specified sheets
gs1_file_path = os.getenv('path_datamodel_GS1')
datamodel_file_path = os.getenv('path_datamodel_maxeda')
CDB_MDM_mapping_path = os.getenv('path_CDB_MDM_mapping')
CBD_GS1_mapping_path = os.getenv('path_CDB_GS1_mapping')
file_path_workflowSKUs = os.getenv('file_path_workflowSKUs')
# Get the output folder path from the environment variables
output_folder = 'Output'

###################
## GS1 datamodel
###################
print('  ### Read GS1 datamodel ###')
# Read the 'Bricks' sheet from the GS1 file, starting from row 4 for headers
gs1_df = pd.read_excel(gs1_file_path, sheet_name='Bricks', skiprows=3, dtype=str)
# Select relevant columns
gs1_df = gs1_df[['Brick Code','Brick activated', 'Brick Title', 'Segment Code', 'Segment Title', 'Family Code', 'Family Title', 'Class Code', 'Class Title', 'FR Brick Title', 'FR Segment Title', 'FR Family Title', 'FR Class Title','NL Brick Title', 'NL Segment Title', 'NL Family Title', 'NL Class Title']].astype(str)

print('## Attributes per brick ##')
# Read the Attributes per Brick sheet from the GS1 file to be able to address the attributes from active bricks
gs1_df_attributes_brick = pd.read_excel(gs1_file_path, sheet_name='Data for Attributes per Brick', skiprows=3, dtype=str)

print('## Read attribute metadata ##')
# Read metadata for attributes
gs1_df_attributes = pd.read_excel(gs1_file_path, sheet_name='Fielddefinitions', skiprows=3, dtype=str)

print('## Read picklist values ##')
# Read picklists
gs1_df_picklists = pd.read_excel(gs1_file_path, sheet_name='Picklists', skiprows=3, dtype=str)

###################
## Maxeda datamodel
###################
print(f'\n  ### Read Maxeda datamodel ###')
def maxeda_sheet(sheet):
    print(f'## Read {sheet} ##')
    # Read the 'S9 - Category' sheet from the Datamodel file
    maxeda_sheet = pd.read_excel(datamodel_file_path, sheet_name=sheet, dtype=str)
    # Select relevant columns
    maxeda_sheet = maxeda_sheet.astype(str).apply(lambda x: x.str.strip("'\""))

    # Apply the function to extract Segment, Family, and Class Codes from the datamodel_bricks
    # print('# Extract parents #')
    maxeda_sheet[['Segment Code', 'Family Code', 'Class Code']] = maxeda_sheet['Parent Category Path'].apply(extract_parents).apply(pd.Series)

    return maxeda_sheet

maxeda_s9 = maxeda_sheet('S9 - Category')
maxeda_s10 = maxeda_sheet('S10 - Category - Locale')

print(f'## Read S7 - Attribute ##')
# Read the 'S7 - Attribute' sheet from the Datamodel file
maxeda_s7_full_df = pd.read_excel(datamodel_file_path, sheet_name='S7 - Attribute')

print(f'## Read S8 - Attribute - Locale ##')
# S8 - Attribute - Locale
maxeda_s8_full_df = pd.read_excel(datamodel_file_path, sheet_name='S8 - Attribute - Locale')

print(f'## Read S23 - Lookup Model ##')
maxeda_s23_df = pd.read_excel(datamodel_file_path, sheet_name='S23 - Lookup Model')

print(f'## Read S14 - CAT - ATTR ##')
maxeda_s14_df = pd.read_excel(datamodel_file_path, sheet_name='S14 - CAT - ATTR')
# Convert 'Category Name' to string (object) in maxeda_s14_df
maxeda_s14_df['Category Name'] = maxeda_s14_df['Category Name'].astype(str)
# print("maxeda_s14_df A:", maxeda_s14_df['Category Name'].dtype)

print(f'## Read CDB_MDM_AttributeMapping ##')
CDB_MDM_mapping_df = pd.read_excel(CDB_MDM_mapping_path, sheet_name='CDB_MDMAttributeMapping')

print(f'## Read CDB_GS1_AttributeMapping ##')
CDB_GS1_mapping_df = pd.read_excel(CBD_GS1_mapping_path, sheet_name='CDB_GS1Attributes')

# Workflow
print(f'## Read Workflow ##')
workflowSKUs_df = pd.read_excel(file_path_workflowSKUs, header=1)  # Header is in the second row

# Selecting required columns from the main DataFrame
workflowSKUs_df_selected = workflowSKUs_df[['Brick', 'GTIN', 'VendorNumberSAP', 'ArticleLongName']]

# print(workflowSKUs_df_selected)
# exit()

# print("workflowSKUs_df_selected A:", workflowSKUs_df_selected['Brick'].dtype)
print('\n  ### Establish active bricks ###')
gs1_df_active= gs1_df[gs1_df['Brick activated'] == 'Yes']
gs1_active_brick_set = set(gs1_df_active['Brick Code'].dropna())
gs1_active_segment_set = set(gs1_df_active['Segment Code'].dropna())
gs1_active_family_set = set(gs1_df_active['Family Code'].dropna())
gs1_active_class_set = set(gs1_df_active['Class Code'].dropna())


def sheet(sheet_df):
    # Define the columns to keep
    if sheet_df.equals(maxeda_s9):
        columns_to_keep = ['ID', 'Action', 'Unique Identifier', 'Category Name', 'Category Long Name', 'Parent Category Path', 'Hierarchy Name']
        sheet_being_processes = 'S9 - Category'
    elif sheet_df.equals(maxeda_s10):
        columns_to_keep = ['ID', 'Action', 'Unique Identifier', 'Category Name', 'Parent Category Path', 'Hierarchy Name', 'Locale', 'Category Long Name']
        sheet_being_processes = 'S10'

    print(f'\n  ### Process {sheet_being_processes} ###')

    # Make a placeholder for all processed items
    all_categories = []

    # Filter the DataFrame to include only rows where 'Hierarchy Name' is equal to 'GS1 Hierarchy'
    base_fiter_datamodel = sheet_df[(sheet_df['Hierarchy Name'].isin(['GS1 Hierarchy']))]

    # Convert datamodel to sets
    filtered_datamodel_brick = base_fiter_datamodel[(base_fiter_datamodel['Parent Category Path'].str.count('//') == 3)]
    # Filter out rows where 'Category Name' starts with '999'
    filtered_datamodel_brick = filtered_datamodel_brick[~filtered_datamodel_brick['Category Name'].str.startswith('999')]

    datamodel_brick_set = set(filtered_datamodel_brick['Category Name'].dropna())
    
    filtered_datamodel_segment = base_fiter_datamodel[(base_fiter_datamodel['Parent Category Path'].str.count('//') == 0)]
    datamodel_segment_set = set(filtered_datamodel_segment['Category Name'].dropna())

    filtered_datamodel_family = base_fiter_datamodel[(base_fiter_datamodel['Parent Category Path'].str.count('//') == 1)]
    datamodel_family_set = set(filtered_datamodel_family['Category Name'].dropna())

    filtered_datamodel_class = base_fiter_datamodel[(base_fiter_datamodel['Parent Category Path'].str.count('//') == 2)]
    datamodel_class_set = set(filtered_datamodel_class['Category Name'].dropna())

    #####################
    ## Add
    #####################
    # Create a list to store data for new categories
    print(f'## Additions - {sheet_being_processes} ##')
    levels_and_data_new = [
        ('brick', gs1_active_brick_set - datamodel_brick_set),
        ('segment', gs1_active_segment_set - datamodel_segment_set),
        ('family', gs1_active_family_set - datamodel_family_set),
        ('class', gs1_active_class_set - datamodel_class_set)
    ]

    # Dictionary to hold individual DataFrames for each level
    individual_new_dfs = {}   

    # Loop through each level to generate corresponding DataFrames
    for level, selection in levels_and_data_new:
        new_categories = create_category_df(selection, level, sheet_df)
        new_categories['Reason'] = 'add: ' + level
        new_categories = new_categories[columns_to_keep]
        all_categories.append(new_categories)

        # Store the individual DataFrame in the dictionary
        individual_new_dfs[level] = new_categories

    # Extract individual DataFrames for each level
    new_brick_categories_df = individual_new_dfs.get('brick')
    new_segment_categories_df = individual_new_dfs.get('segment')
    new_family_categories_df = individual_new_dfs.get('family')
    new_class_categories_df = individual_new_dfs.get('class')

    #####################
    ## Workflow
    #####################
    print(f'## Process workflow SKUs {sheet_being_processes} ##')

    # Select only the columns needed for the merge from sheet
    sheet_relevant_columns = sheet_df[['Category Name', 'Segment Code', 'Family Code', 'Class Code']]
   
    # Merge with gs1_file_path_workflowSKUs_df to get the Segment Code, Family Code, and Class Code
    workflowSKUs_df_merged = pd.merge(workflowSKUs_df_selected, sheet_relevant_columns, how='left', left_on='Brick', right_on='Category Name')

    def workflow_aggregation(level):
        if level not in ['Brick', 'Segment Code', 'Family Code', 'Class Code']:
            raise ValueError("Invalid aggregation level")

        # Group by the specified level and aggregate
        result = workflowSKUs_df_merged.groupby(level).agg({
            'GTIN': 'nunique',  # Count of unique SKUs
            'VendorNumberSAP': 'nunique', # Unique Vendors
            'ArticleLongName': 'nunique'  # Unique count of ArticleLongName
        }).reset_index()  # Reset index to make the level a column again

        # Rename columns
        result.columns = ['Category', 'unique_count_GTIN', 'unique_count_VendorNumberSAP', 'unique_count_ArticleLongName']

        # Add a new column to denote the level
        result['Level'] = level

        return result

    # Aggregate for each level
    brick_workflow_agg = workflow_aggregation('Brick')
    segment_workflow_agg = workflow_aggregation('Segment Code')
    family_workflow_agg = workflow_aggregation('Family Code')
    class_workflow_agg = workflow_aggregation('Class Code')

    # Print the resulting DataFrames
    # print("Brick Aggregation:")
    # print(brick_workflow_agg)
    # print("\nSegment Aggregation:")
    # print(segment_workflow_agg)
    # print("\nFamily Aggregation:")
    # print(family_workflow_agg)
    # print("\nClass Aggregation:")
    # print(class_workflow_agg)
 
    # We only delete bricks if they are not in the model at all NOT ONLY looking at active bricks
    gs1_brick_set = set(gs1_df['Brick Code'].dropna())
    gs1_segment_set = set(gs1_df['Segment Code'].dropna())
    gs1_family_set = set(gs1_df['Family Code'].dropna())
    gs1_class_set = set(gs1_df['Class Code'].dropna())

    #####################
    ## Delete and backlog
    #####################
    print(f'## Deletions - {sheet_being_processes} ##')
 
    def configure_deletions(datamodel_set, gs1_set, workflow_agg, category_name):
        """
        Configure deletions for a given category.

        Args:
        datamodel_set (set): The set of categories in the data model.
        gs1_set (set): The set of categories in the GS1 data.
        workflow_agg (pd.DataFrame): The workflow aggregation DataFrame for the category.
        category_name (str): The name of the category (e.g., 'Brick', 'Segment', 'Family', 'Class').

        Returns:
        pd.DataFrame: DataFrame of backlog deletions for the category.
        """
        initial_deletion_set = datamodel_set - gs1_set
        workflow_set = set(workflow_agg['Category'].dropna())
        final_deletion_set = initial_deletion_set - workflow_set
        backlog_deletion_set = initial_deletion_set & workflow_set

        backlog_deletions = workflow_agg[workflow_agg['Category'].isin(backlog_deletion_set)].copy()
        if not backlog_deletions.empty:
            backlog_deletions.loc[:, 'Reason'] = f'{category_name} deletion'

        return workflow_set, final_deletion_set, backlog_deletions

    # Configure deletions for each category
    brick_workflow_set, final_brick_deletion_set, backlog_brick_deletions = configure_deletions(datamodel_brick_set, gs1_brick_set, brick_workflow_agg, 'Brick')
    segment_workflow_set, final_segment_deletion_set, backlog_segment_deletions = configure_deletions(datamodel_segment_set, gs1_segment_set, segment_workflow_agg, 'Segment')
    family_workflow_set, final_family_deletion_set, backlog_family_deletions = configure_deletions(datamodel_family_set, gs1_family_set, family_workflow_agg, 'Family')
    class_workflow_set, final_class_deletion_set, backlog_class_deletions = configure_deletions(datamodel_class_set, gs1_class_set, class_workflow_agg, 'Class')

    # print(f"initial_brick_deletion_set: {initial_brick_deletion_set}")
    # print(f"workflow_brick_set: {set(brick_workflow_agg['Brick'])}")
    # print(f"backlog_brick_deletion_set: {backlog_brick_deletion_set}")
    # print(f"backlog_brick_deletions: {backlog_brick_deletions}")

    # Create a list to store data for new categories
    levels_and_data_delete = [
        ('brick', final_brick_deletion_set),
        ('segment', final_segment_deletion_set),
        ('family', final_family_deletion_set),
        ('class', final_class_deletion_set)
    ]

    # Dictionary to hold individual DataFrames for each level
    individual_delete_dfs = {}
    
    # Loop through the levels and update the combined set with each difference set
    for level, difference_set in levels_and_data_delete:
        delete_categories = delete_category_df(difference_set, sheet_df)

        # delete_categories['Reason'] = 'Delete'
        delete_categories = delete_categories[columns_to_keep]
        all_categories.append(delete_categories)

        # Store the individual DataFrame in the dictionary
        individual_delete_dfs[level] = delete_categories

    # Extract individual DataFrames for each level
    delete_brick_categories_df = individual_delete_dfs.get('brick')
    delete_segment_categories_df = individual_delete_dfs.get('segment')
    delete_family_categories_df = individual_delete_dfs.get('family')
    delete_class_categories_df = individual_delete_dfs.get('class')

    delete_segment_categories_df = delete_segment_categories_df[delete_segment_categories_df['Category Long Name'] != 'GS1 - GS1']

    ######################
    ## Change in hierachy of active bricks & Backlog
    #####################
    print(f'## Hierarchy change - {sheet_being_processes} ##')

    def merge_and_check_hierarchy(gs1_df, sheet_df, merge_columns, check_columns, hierarchy_name):
        # Merge the two dataframes on the specified columns
        left_on, right_on = merge_columns
        comparison_df = pd.merge(gs1_df, sheet_df, how='inner', left_on=left_on, right_on=right_on)

        # Check for mismatched parents using the provided columns
        condition = False
        for col_x, col_y in check_columns:
            condition |= (comparison_df[f'{col_x}_x'] != comparison_df[f'{col_y}_y'])

        change_hierarchy_df = comparison_df[condition].copy()

        # Create the new 'Parent Category Path' based on the given conditions
        change_hierarchy_df['Parent Category Path'] = change_hierarchy_df.apply(
            lambda row: f"GS1//{row[f'{check_columns[0][0]}_x']}//{'//'.join([row[f'{col[0]}_x'] for col in check_columns[1:]])}" 
            if row['Hierarchy Name'] == hierarchy_name
            else f"{row[f'{check_columns[0][0]}_x']}//{'//'.join([row[f'{col[0]}_x'] for col in check_columns[1:]])}",
            axis=1
        )

        # Replace 'nan' strings with an empty string in specific columns
        change_hierarchy_df['Action'] = change_hierarchy_df['Action'].replace('nan', '')
        change_hierarchy_df['Unique Identifier'] = change_hierarchy_df['Unique Identifier'].replace('nan', '')

        change_hierarchy_df = change_hierarchy_df[columns_to_keep]
        
        return change_hierarchy_df

    # Specify the hierarchy name
    hierarchy_name = 'GS1 Hierarchy'

    # Define the check columns for each level
    brick_check_columns = [('Segment Code', 'Segment Code'), ('Family Code', 'Family Code'), ('Class Code', 'Class Code')]
    class_check_columns = [('Segment Code', 'Segment Code'), ('Family Code', 'Family Code')]
    family_check_columns = [('Segment Code', 'Segment Code')]

    # Process each hierarchy level
    # For bricks, the merge columns are different
    change_hierarchy_brick_df = merge_and_check_hierarchy(gs1_df_active, sheet_df, ('Brick Code', 'Category Name'), brick_check_columns, hierarchy_name)
    change_hierarchy_class_df = merge_and_check_hierarchy(gs1_df_active, sheet_df, ('Class Code', 'Category Name'), class_check_columns, hierarchy_name)
    change_hierarchy_family_df = merge_and_check_hierarchy(gs1_df_active, sheet_df, ('Family Code', 'Category Name'), family_check_columns, hierarchy_name)

        #####################
        ## Integrate backlog
        #####################
    
    def configure_changes(change_df, workflow_set, workflow_agg, reason):
        initial_change_set = set(change_df['Category Name'].dropna())
        final_change_set = initial_change_set - workflow_set
        backlog_change_set = initial_change_set & workflow_set

        backlog_changes = workflow_agg[workflow_agg['Category'].isin(backlog_change_set)].copy()
        if not backlog_changes.empty:
            backlog_changes.loc[:, 'Reason'] = reason

        return final_change_set, backlog_changes

    final_brick_change_set, backlog_brick_changes = configure_changes(change_hierarchy_brick_df, brick_workflow_set, brick_workflow_agg, 'Brick hierarchy change')
    final_family_change_set, backlog_family_changes = configure_changes(change_hierarchy_family_df, family_workflow_set, family_workflow_agg, 'Family hierarchy change')
    final_class_change_set, backlog_class_changes = configure_changes(change_hierarchy_class_df, class_workflow_set, class_workflow_agg, 'Class hierarchy change')

    change_hierarchy_brick_df = change_hierarchy_brick_df[change_hierarchy_brick_df['Category Name'].isin(final_brick_change_set)]
    change_hierarchy_class_df = change_hierarchy_class_df[change_hierarchy_class_df['Category Name'].isin(final_class_change_set)]
    change_hierarchy_family_df = change_hierarchy_family_df[change_hierarchy_family_df['Category Name'].isin(final_family_change_set)]


    # Append the results to the all_categories list
    all_categories.append(change_hierarchy_brick_df)
    all_categories.append(change_hierarchy_class_df)
    all_categories.append(change_hierarchy_family_df)

 
    #####################
    ## Combine backlogs
    #####################

   # List of DataFrames to concatenate
    backlog_dfs = [backlog_brick_deletions, backlog_segment_deletions, backlog_family_deletions, backlog_class_deletions, backlog_brick_changes, backlog_family_changes, backlog_class_changes]

    # Concatenate all DataFrames into a single DataFrame
    backlog_df = pd.concat(backlog_dfs, ignore_index=True)

    #####################
    ## Combine all categories for full output
    #####################

    # Concatenate all DataFrames into one, if there are any DataFrames to concatenate
    if all_categories:
        final_all_categories = pd.concat(all_categories, ignore_index=True)
    else:
        final_all_categories = pd.DataFrame()  # Fallback to an empty DataFrame if no data

    #####################
    ## Return both full output and subsections
    #####################
    return final_all_categories, delete_brick_categories_df, delete_class_categories_df, delete_family_categories_df, delete_segment_categories_df, new_brick_categories_df, new_class_categories_df, new_family_categories_df, new_segment_categories_df, change_hierarchy_brick_df, change_hierarchy_class_df, change_hierarchy_family_df, backlog_df, brick_workflow_set

final_all_categories_s9, delete_brick_categories_s9, delete_class_categories_s9, delete_family_categories_s9, delete_segment_categories_s9, new_brick_categories_s9, new_class_categories_s9, new_family_categories_s9, new_segment_categories_s9, change_hierarchy_bricks_s9, change_hierarchy_classes_s9, change_hierarchy_families_s9, backlog_S9_df, brick_workflow_set_S9 = sheet(maxeda_s9)
final_all_categories_locale_combined_s10, delete_brick_categories_s10, delete_class_categories_s10, delete_family_categories_s10, delete_segment_categories_s10, new_brick_categories_s10, new_class_categories_s10, new_family_categories_s10, new_segment_categories_s10, change_hierarchy_bricks_s10, change_hierarchy_classes_s10, change_hierarchy_families_s10, backlog_S10_df, brick_workflow_set_S10 = sheet(maxeda_s10)

final_backlog_df = pd.concat([backlog_S9_df, backlog_S10_df], ignore_index=True)
final_backlog_df = final_backlog_df.drop_duplicates()

# print(final_backlog_df)
# exit()

# Combine the sets and remove duplicates
final_brick_workflow_set = brick_workflow_set_S9 | brick_workflow_set_S10

########################################################
## Attributes
########################################################

# 1 load GS1 & Maxeda datamodels
# 2 perform changes on picklistvalues (inherent with deletions) so that picklist-attributes can be delete later on
# 3 additions, deletions and changes for S7, S8, S23 (S7 can also change)
# 4 additions to picklistvalues for new picklists and attributes changed to picklist) 

###################
## GS1 datamodel
###################

print('\n  ### Establish attributes from active brick ###\n')
# Filter gs1_df_attributes_brick for only those rows where the 'Brick' column's values are in gs1_active_brick_set
gs1_df_attributes_brick_active = gs1_df_attributes_brick[gs1_df_attributes_brick['Brick'].isin(gs1_active_brick_set)]

# Create a set of the 'FieldID' values from the filtered DataFrame
gs1_attributes_GS1ID_set = set(gs1_df_attributes_brick_active['FieldID'].dropna())

# Check if 'Attributename English' column has the value 'Battery size type code'
mask = gs1_df_attributes_brick_active['FieldID'] == '8.384'

# # Check if 'BatteryTypeCode' is in the set
# if '8.384' in gs1_attributes_GS1ID_set:
#     print("'BatteryTypeCode' is in the set")
# else:
#     print("'BatteryTypeCode' is not in the set")

# exit()    


###################
## Maxeda datamodel
###################

# Extract attribute code after "id" (case-insenstive)
def extract_attribute_code(definition):
    # Convert to string in case the input is NaN or any other non-string data
    definition = str(definition)
    pattern = re.compile(r"GS1 Field[_ ]ID (\S+)|GS1 FieldID(\S+)", re.IGNORECASE)
    match = pattern.search(definition)
    if match:
        return match.group(1).strip()  # Extract and return the part after 'id '
    return ''  # Return an empty string if no match is found


# Select scope
maxeda_s7_df_scope = maxeda_s7_full_df[maxeda_s7_full_df['Attribute Type'].isin(['Category', 'Common'])].copy()
# maxeda_s7_df_scope = maxeda_s7_full_df[maxeda_s7_full_df['Attribute Type'].isin(['Category'])].copy()

# Extract Attribute codes
maxeda_s7_df_scope['Attribute code'] = maxeda_s7_df_scope['Definition'].apply(extract_attribute_code)

# Exclude maxeda-attributes
maxeda_s7_df_scope = maxeda_s7_df_scope[~maxeda_s7_df_scope['Attribute code'].str.startswith("M")]

# Convert the 'Precision' column to string type
maxeda_s7_df_scope['Precision'] = maxeda_s7_df_scope['Precision'].astype(str)
# Use string manipulation to remove trailing '.0' from the string representation
maxeda_s7_df_scope['Precision'] = maxeda_s7_df_scope['Precision'].str.replace(r'\.0$', '', regex=True)
# Replace 'nan' strings with empty strings
maxeda_s7_df_scope['Precision'] = maxeda_s7_df_scope['Precision'].replace('nan', '')

# # Check if 'Attributename English' column has the value 'Battery size type code'
# mask = maxeda_s7_df_scope['Attribute code'] == '8.384'

# # Print 'Picklist ID' for rows where the condition is true
# if mask.any():
#     print(maxeda_s7_df_scope[mask]['Attribute code'].values[0])

# exit()

# Select scope
maxeda_s8_df_scope = maxeda_s8_full_df[maxeda_s8_full_df['Attribute Path'].str.contains('Category Specific Attributes//|OneWS_.*OneWS_', na=False)].copy()

# Convert the 'ID' column of maxeda_s8_df to a set
maxeda_s8_IDs_set = set(maxeda_s8_df_scope['ID'])


maxeda_s8_df_scope['Attribute code'] = maxeda_s8_df_scope['Definition'].apply(extract_attribute_code)
# Correct for the fact that not every attributeID is formatted with a dot on the second position
maxeda_s8_df_scope['Attribute code'] = maxeda_s8_df_scope['Attribute code'].apply(
    lambda x: x[0] + '.' + x[1:] if len(x) > 1 and x[1] != '.' else x
)

# Exclude maxeda-attributes
maxeda_s8_df_scope = maxeda_s8_df_scope[~maxeda_s8_df_scope['Attribute code'].str.startswith("M")]


#####################################
## Workflow
#####################################

# print("test 1")
brick_attributes_workflow_df = pd.merge(
    workflowSKUs_df_selected[['Brick', 'GTIN', 'ArticleLongName', 'VendorNumberSAP']],
    maxeda_s14_df[['Category Name', 'Attribute Path']],
    how='inner',
    left_on='Brick',
    right_on='Category Name'
)


# print(f'brick_attributes_workflow_df: {brick_attributes_workflow_df.head}')

# print("test 2")
# De-duplicate the mappping between Attribute Path and ID to be able to effectively join upon
maxeda_s8_AttributePath_ID_mapping = maxeda_s8_full_df[['Attribute Path', 'ID']].copy()
maxeda_s8_AttributePath_ID_mapping = maxeda_s8_AttributePath_ID_mapping.drop_duplicates(subset='ID')

# print(f'brick_attributes_workflow_df2: {maxeda_s8_AttributePath_ID_mapping.head}')

# print("test 3")
# Perform the left join on 'Attribute Path' and select only the required columns
brick_attributes_workflow_df2 = pd.merge(
    brick_attributes_workflow_df, 
    maxeda_s8_AttributePath_ID_mapping, 
    how='left', 
    on='Attribute Path'
)
# print(f'brick_attributes_workflow_df2: {brick_attributes_workflow_df2.head}')

# print("test 4")
brick_attributes_workflow_df3 = pd.merge(
    brick_attributes_workflow_df2, 
    maxeda_s7_df_scope[['ID', 'Attribute code']], 
    how='left', 
    on='ID'
)
# print(f'brick_attributes_workflow_df3: {brick_attributes_workflow_df3.head}')

brick_attributes_workflow_df4 = pd.merge(
    brick_attributes_workflow_df3, 
    maxeda_s7_df_scope[['LookUp Table Name', 'Attribute code']], 
    how='left', 
    on='Attribute code'
)

# exit()

# print("test 5")
brick_attributes_aggregation_workflow_df_temp = brick_attributes_workflow_df3.groupby(['Brick', 'Attribute code']).agg({
    'GTIN': pd.Series.nunique,
    'VendorNumberSAP': pd.Series.nunique,
    'ArticleLongName': pd.Series.nunique    
}).reset_index()

# Prefix all headers with 'unique_count_' for brick_attributes_aggregation_workflow_df
brick_attributes_aggregation_workflow_df_temp.columns = ['unique_count_' + col if col not in ['Brick', 'Attribute code'] else col for col in brick_attributes_aggregation_workflow_df_temp.columns]

# print(brick_attributes_aggregation_workflow_df_temp.head)
# print(brick_attributes_workflow_df3.head)

# exit()

maxeda_s8_AttributePath_AttributeCode_mapping = brick_attributes_workflow_df3[['Attribute Path', 'Attribute code']].copy()
maxeda_s8_AttributePath_AttributeCode_mapping = maxeda_s8_AttributePath_AttributeCode_mapping.drop_duplicates(subset='Attribute code')

brick_attributes_aggregation_workflow_df = pd.merge(
    brick_attributes_aggregation_workflow_df_temp, 
    maxeda_s8_AttributePath_AttributeCode_mapping, 
    how='left', 
    on='Attribute code'
)

# print("test 6")
attributes_aggregation_workflow_GS1_ID_df = brick_attributes_workflow_df3.groupby(['Attribute code']).agg({
    'GTIN': pd.Series.nunique,
    'VendorNumberSAP': pd.Series.nunique,
    'ArticleLongName': pd.Series.nunique    
}).reset_index()

# Prefix all headers with 'unique_count_' for attributes_aggregation_workflow_GS1_ID_df
attributes_aggregation_workflow_GS1_ID_df.columns = ['unique_count_' + col if col != 'Attribute code' else col for col in attributes_aggregation_workflow_GS1_ID_df.columns]

# print("test 7")
attributes_aggregation_workflow_ALL_ID_df = pd.merge(
    attributes_aggregation_workflow_GS1_ID_df, 
    maxeda_s7_df_scope[['ID', 'Attribute Name', 'Attribute code']], 
    how='left', 
    on='Attribute code'
)

attributes_aggregation_workflow_lookuptable_df = brick_attributes_workflow_df4.groupby(['LookUp Table Name']).agg({
    'GTIN': pd.Series.nunique,
    'VendorNumberSAP': pd.Series.nunique,
    'ArticleLongName': pd.Series.nunique    
}).reset_index()

# Prefix all headers with 'unique_count_' for attributes_aggregation_workflow_GS1_ID_df
attributes_aggregation_workflow_lookuptable_df.columns = ['unique_count_' + col if col != 'LookUp Table Name' else col for col in attributes_aggregation_workflow_lookuptable_df.columns]


maxeda_s23_df_selected = maxeda_s23_df[['ID', 'Table Name']]
maxeda_s23_df_selected = maxeda_s23_df_selected.drop_duplicates()   


attributes_aggregation_workflow_lookuptable_df = pd.merge(
    attributes_aggregation_workflow_lookuptable_df, 
    maxeda_s23_df_selected[['ID', 'Table Name']], 
    how='left', 
    left_on='LookUp Table Name',
    right_on='Table Name',
)

# print(f'attributes_aggregation_workflow_ALL_ID_df: {attributes_aggregation_workflow_ALL_ID_df.head}')


# print(attributes_aggregation_workflow_ALL_ID_df.columns)
# print(attributes_aggregation_workflow_ALL_ID_df.head)


# exit()


# print("test")
# # Group by 'Brick' and count unique values of 'GTIN', 'ArticleLongName', and 'VendorNumberSAP'
# brick_workflow_total_df = workflowSKUs_df.groupby('Brick').agg({
#     'GTIN': pd.Series.nunique,
#     'VendorNumberSAP': pd.Series.nunique,
#     'ArticleLongName': pd.Series.nunique    
# }).reset_index()



# # Rename columns for clarity
# brick_aggregation_workflow_total_df.columns = ['Brick', 'Unique_GTIN_Count', 'Unique_VendorNumberSAP', 'Unique_ArticleLongName_Count']

# # brick_workflow_total_df = workflowSKUs_df.copy()
# print("test1")

# brick_aggregation_workflow_with_attributes_df_temp = pd.merge(
#     brick_aggregation_workflow_total_df,
#     maxeda_s14_df[['Category Name', 'Attribute Path']],
#     how='inner',
#     left_on='Brick',
#     right_on='Category Name'
# )
# brick_aggregation_workflow_with_attributes_df_temp.drop(columns=['Category Name'], inplace=True)
# # print(brick_workflow_with_attributes_df.columns)
# print("test2")

# # Perform the left join on 'Attribute Path' and select only the required columns
# brick_aggregation_workflow_with_attributes_df_temp2 = pd.merge(
#     brick_aggregation_workflow_with_attributes_df_temp, 
#     maxeda_s8_full_df[['Attribute Path', 'ID']], 
#     how='left', 
#     on='Attribute Path'
# )

# brick_aggregation_workflow_with_attributes_df_temp2 = brick_aggregation_workflow_with_attributes_df_temp2.drop_duplicates(subset=['Attribute Path', 'ID', 'Brick'])

# brick_aggregation_workflow_with_attributes_df = pd.merge(
#     brick_aggregation_workflow_with_attributes_df_temp2, 
#     maxeda_s7_df_scope[['ID', 'Attribute code']], 
#     how='left', 
#     on='ID'
# )

# # Group by 'ID' and sum the relevant columns
# MaxedaAttributeID_aggregation_workflow_df = brick_aggregation_workflow_with_attributes_df.groupby(['ID', 'Attribute code']).agg({
#     'Unique_GTIN_Count': 'sum',
#     'Unique_VendorNumberSAP': 'sum',
#     'Unique_ArticleLongName_Count': 'sum'
# }).reset_index()
# print("test3")


# Attribute_allIDs_aggregation_workflow_df = pd.merge(
#     MaxedaAttributeID_aggregation_workflow_df, 
#     maxeda_s7_df_scope[['ID','LookUp Table Name', 'Attribute Name', 'Attribute code']], 
#     how='left', 
#     on='ID'
# )


####################
## Pre-calculations for possible additions based on S7, catspec
####################
print(f'\n  ### Pre-calculations for possible additions ###')

# Get relevant rows from attribute overview
gs1_df_attributes_processed = gs1_df_attributes.copy()

# Data Type and display type
def determine_types(row):
    format = row['Format']
    decimals = row['Deci-\nmals']
    if format == "Number":
        data_type = "Integer" if decimals == '0' else "Decimal"
        display_type = "NumericTextBox"
    elif format == "DateTime":
        data_type = "DateTime"
        display_type = "DateTime"
    elif format == "Text":
        data_type = "String"
        display_type = "TextBox" 
    elif format == "Picklist (T/F)":
        data_type = "String"
        display_type = "LookupTable"
    elif format == "Picklist":
        data_type = "String"
        display_type = "LookupTable"
    elif format == "NumberPicklist":
        data_type = "Integer" if decimals == '0' else "Decimal"
        display_type = "NumericTextBox"
    elif format == "Boolean":
        data_type = "String"
        display_type = "LookupTable"
    else:
        data_type = "Unknown"
        display_type = "Unknown"
    return pd.Series([data_type, display_type], index=['INPUT_Data_type', 'INPUT_Display_type'])

# Apply the function to each row of the dataframe
gs1_df_attributes_processed[['INPUT_Data_type', 'INPUT_Display_type']] = gs1_df_attributes_processed.apply(determine_types, axis=1)

# 
gs1_df_attributes_processed['INPUT_Attribute_name'] = gs1_df_attributes_processed['Attributename English'].apply(
                                                lambda x: x[:x.rfind('(')].strip() if '(' in x and x.endswith(')') else x.strip()
                                            ).apply(lambda x: f"CatSpec_{x}")


def clean_attribute_name(sheet_name):
    # Define invalid characters for Excel sheet names
    invalid_chars = ['\\', '/', '*', '[', ']', ':', '?', ' ']
    # Remove invalid characters
    for char in invalid_chars:
        sheet_name = sheet_name.replace(char, '')

    return sheet_name

# Apply the function to the INPUT_Attribute_name column
gs1_df_attributes_processed['INPUT_Attribute_name'] = gs1_df_attributes_processed['INPUT_Attribute_name'].apply(clean_attribute_name)

# INPUT_Lookup_table_name
gs1_df_attributes_processed['INPUT_Lookup_table_name'] = np.select(
    [
        gs1_df_attributes_processed['Format'].isin(["Picklist (T/F)", "Boolean"]),
        gs1_df_attributes_processed['Format'] == "Picklist"
    ],
    [
        "YesNo",
        gs1_df_attributes_processed['INPUT_Attribute_name'].str.replace(r'\s+', '', regex=True).str.strip().apply(lambda x: x[8:][:30]) # max is 30 charcters for an Excel sheet, but need space for OneWs later on
    ],
    default=""
)

# Group all allowed UOM's INPUT_Allowed_uoms
code_value_concat = gs1_df_picklists.groupby('Picklist ID')['Code value'].apply(lambda x: '||'.join(x.dropna().astype(str))).rename('INPUT_Allowed_uoms')

# Using a left join ensures all original rows in gs1_df_attributes_processed are retained
gs1_df_attributes_processed = gs1_df_attributes_processed.merge(code_value_concat, on='Picklist ID', how='left')

# Fill NaNs with empty strings if any picklist IDs didn't have code values
gs1_df_attributes_processed['INPUT_Allowed_uoms'] = gs1_df_attributes_processed['INPUT_Allowed_uoms'].fillna('')


#Fill the table for S7 as default
gs1_df_attributes_processed['ID'] = ''
gs1_df_attributes_processed['Action'] = ''
gs1_df_attributes_processed['Unique Identifier'] = ''
gs1_df_attributes_processed['Attribute Type'] = 'Category'
gs1_df_attributes_processed['Attribute Name'] = gs1_df_attributes_processed['INPUT_Attribute_name']
gs1_df_attributes_processed['Attribute Long Name'] = gs1_df_attributes_processed['Attributename English']
gs1_df_attributes_processed['Attribute Parent Name'] = 'Category Specific Attributes'
gs1_df_attributes_processed['Data Type'] = gs1_df_attributes_processed['INPUT_Data_type']
gs1_df_attributes_processed['Display Type'] = gs1_df_attributes_processed['INPUT_Display_type']
gs1_df_attributes_processed['Is Collection'] = np.where(
                                            gs1_df_attributes_processed['Repeat'].str.len() > 0, 'YES', 'NO'
                                        )
gs1_df_attributes_processed['Is Inheritable'] = 'NO' #For OneWS is 'YES' bij picklisten en numberpicklisten
gs1_df_attributes_processed['Is Localizable'] = 'NO'
gs1_df_attributes_processed['Is Complex'] = 'NO'
gs1_df_attributes_processed['Is Lookup'] =  np.where(gs1_df_attributes_processed['INPUT_Display_type'] == 'LookupTable', 'YES', 'NO')
gs1_df_attributes_processed['Is Required'] = 'NO'
gs1_df_attributes_processed['Is ReadOnly'] = 'NO'
gs1_df_attributes_processed['Is Hidden'] = 'NO'
gs1_df_attributes_processed['Show At Entity Creation?'] = 'YES'
gs1_df_attributes_processed['Is Searchable'] = 'YES'
gs1_df_attributes_processed['Is Null Value Search Required'] = 'YES'
gs1_df_attributes_processed['Generate Report Table Column?'] = ''
gs1_df_attributes_processed['Default Value'] = ''
gs1_df_attributes_processed['Minimum Length'] = 0  
gs1_df_attributes_processed['Maximum Length'] = 0  
gs1_df_attributes_processed['Range From'] = ''
gs1_df_attributes_processed['Is Range From Inclusive'] = ''
gs1_df_attributes_processed['Range To'] = ''
gs1_df_attributes_processed['Is Range To Inclusive'] = ''
gs1_df_attributes_processed['Precision'] = gs1_df_attributes_processed['Precision'] = gs1_df_attributes_processed['Deci-\nmals'].replace('0', '')
gs1_df_attributes_processed['Use Arbitrary Precision?'] = np.where(
                                                        gs1_df_attributes_processed['Precision'].str.len() > 0, 'NO', ''
                                                    )
gs1_df_attributes_processed['UOM Type'] = np.where(gs1_df_attributes_processed['Format'] == 'NumberPicklist', 'Custom UOM', '') #numberbicklist ? --> "CustomUOM",  bij onews "gdsn uom"
gs1_df_attributes_processed['Allowed UOMs'] = np.where(gs1_df_attributes_processed['Format'] == 'NumberPicklist', gs1_df_attributes_processed['INPUT_Allowed_uoms'],'') #ONLY FOR numberbicklist
gs1_df_attributes_processed['Default UOM'] = np.where(gs1_df_attributes_processed['Format'] == 'NumberPicklist', gs1_df_attributes_processed['UoM fixed'],'')
gs1_df_attributes_processed['Allowable Values'] = ''
gs1_df_attributes_processed['LookUp Table Name'] = gs1_df_attributes_processed['INPUT_Lookup_table_name']
gs1_df_attributes_processed['Lookup Display Columns'] = gs1_df_attributes_processed['LookUp Table Name'].apply(lambda x: f"[{x}]" if x.strip() else x)
gs1_df_attributes_processed['Lookup Search Columns'] = gs1_df_attributes_processed['LookUp Table Name'].apply(lambda x: f"[{x}]" if x.strip() else x)
gs1_df_attributes_processed['Lookup Display Format'] = gs1_df_attributes_processed['LookUp Table Name'].apply(lambda x: f"[{x}]" if x.strip() else x)
gs1_df_attributes_processed['Lookup Sort Order'] = gs1_df_attributes_processed['LookUp Table Name'].apply(lambda x: f"[{x}]" if x.strip() else x)
gs1_df_attributes_processed['Export Format'] = gs1_df_attributes_processed['LookUp Table Name'].apply(lambda x: f"[{x}]" if x.strip() else x)
gs1_df_attributes_processed['Sort Order'] = 0
gs1_df_attributes_processed['Definition'] = ("GS1 Field_ID " + 
                                         gs1_df_attributes_processed['FieldID'].astype(str) + " " + 
                                         gs1_df_attributes_processed['Definition English'])
gs1_df_attributes_processed['Example'] = ''
gs1_df_attributes_processed['Business Rule'] = ''
gs1_df_attributes_processed['Label'] = ''
gs1_df_attributes_processed['Extension'] = ''
gs1_df_attributes_processed['Web URI'] = ''
gs1_df_attributes_processed['Enable History'] = 'YES'
gs1_df_attributes_processed['Apply Time Zone Conversion'] = 'NO'
gs1_df_attributes_processed['Attribute Regular Expression'] = ''
gs1_df_attributes_processed['Is UOM Localizable'] = 'NO'

gs1_df_attributes_processed['ONEWS_Is Localizable'] = np.where(gs1_df_attributes_processed['Multi\nlingual'] == 'Yes', 'YES', 'NO') #numberbicklist ? --> "CustomUOM",  bij onews "gdsn uom"

gs1_df_attributes_processed.fillna('', inplace=True)

# Check if 'Attributename English' column has the value 'Battery size type code'
mask = gs1_df_attributes_processed['Attributename English'] == 'Battery size type code'

# Print 'Picklist ID' for rows where the condition is true
# if mask.any():
#     print(gs1_df_attributes_processed[mask]['Picklist ID'].values[0])

# exit()

####################
## S7 
####################
print(f'\n  ### Process S7 ###')

# Select Category to base changes on
maxeda_s7_df_category = maxeda_s7_df_scope[maxeda_s7_df_scope['Attribute Type'] == 'Category'].copy()
# Create a set from this
maxeda_attribute_s7_GS1ID_set = set(maxeda_s7_df_category['Attribute code'].replace('', np.nan).dropna())


# # Check if 'BatteryTypeCode' is in the set
# if '8.384' in maxeda_attribute_s7_GS1ID_set:
#     print("'BatteryTypeCode' is in the set maxeda_attribute_s7_GS1ID_set")
# else:
#     print("'BatteryTypeCode' is not in the set maxeda_attribute_s7_GS1ID_set")

# exit()

    ####################
    ## Establish set of attributes for 1) additions, 2) deletions, and 3) overlapping 
    ####################

attribute_add_s7_GS1ID_set = gs1_attributes_GS1ID_set - maxeda_attribute_s7_GS1ID_set
attribute_delete_s7_GS1ID_set = maxeda_attribute_s7_GS1ID_set - gs1_attributes_GS1ID_set
attribute_overlap_s7_GS1ID_set = gs1_attributes_GS1ID_set & maxeda_attribute_s7_GS1ID_set

# # Check if 'BatteryTypeCode' is in the set
# if '8.384' in gs1_attributes_GS1ID_set:
#     print("'BatteryTypeCode' is in the set gs1_attributes_GS1ID_set")
# else:
#     print("'BatteryTypeCode' is not in the set gs1_attributes_GS1ID_set")
#     print(gs1_attributes_GS1ID_set)

# # Check if 'BatteryTypeCode' is in the set
# if '8.384' in attribute_add_s7_GS1ID_set:
#     print("'BatteryTypeCode' is in the set attribute_add_s7_GS1ID_set")
# else:
#     print("'BatteryTypeCode' is not in the set attribute_add_s7_GS1ID_set")
#     print(attribute_add_s7_GS1ID_set)

# exit()

    ####################
    ## Delete
    ####################

print(f'## Deletions - S7 ##')

delete_attributes_s7_df_temp = maxeda_s7_df_scope[maxeda_s7_df_scope['Attribute code'].isin(attribute_delete_s7_GS1ID_set)].copy()
delete_attributes_s7_df_temp ['Action'] = 'Delete'

# Filter delete_attributes_s7_df to exclude rows with IDs found in MaxedaAttributeID_aggregation_workflow_df['Category']
delete_attributes_s7_df = delete_attributes_s7_df_temp[~delete_attributes_s7_df_temp['ID'].isin(attributes_aggregation_workflow_ALL_ID_df['ID'])].copy()


# Create backlog items for attribute deletions
    # Find the overlap
backlog_delete_attributes_s7_ID_set = set(delete_attributes_s7_df_temp['ID']).intersection(set(attributes_aggregation_workflow_ALL_ID_df['ID']))
backlog_attribute_deletions = attributes_aggregation_workflow_ALL_ID_df[attributes_aggregation_workflow_ALL_ID_df['ID'].isin(backlog_delete_attributes_s7_ID_set)].copy()
backlog_attribute_deletions['Level'] = 'Attribute'
backlog_attribute_deletions['Reason'] = 'Attribute deletion'

backlog_attribute_deletions['Category'] = (
    backlog_attribute_deletions['Attribute code'] + ' - ' +
    backlog_attribute_deletions['ID'].astype(str) + ' - ' +
    backlog_attribute_deletions['Attribute Name']
)

final_backlog_df = pd.concat([final_backlog_df, backlog_attribute_deletions], ignore_index=True)

# Create sets for re-use later
attribute_delete_s7_MaxedaIDs_set = set(delete_attributes_s7_df['ID'].dropna())
attribute_delete_s7_AttributeName_set = set(delete_attributes_s7_df['Attribute Name'].dropna())

filtered_df = delete_attributes_s7_df_temp[delete_attributes_s7_df_temp['ID'].isin(backlog_delete_attributes_s7_ID_set)]
attribute_delete_s7_AttributeName_backlog_set = set(filtered_df['Attribute Name'].dropna())

# Create lookuptable set as well
attribute_delete_s7_LookupTableName_set = set(delete_attributes_s7_df['LookUp Table Name'].dropna())
attribute_delete_s7_LookupTableName_backlog_set = set(filtered_df['LookUp Table Name'].dropna())

attribute_delete_s7_LookupTableName_set.discard('YesNo')
attribute_delete_s7_LookupTableName_backlog_set.discard('YesNo')


    ####################
    ## Additions
    ####################
print(f'## Additions - S7 ##')
def add_attributes_s7(add_set, all_additions_attributes_s7_df):

    # # Check if 'BatteryTypeCode' is in the set
    # if '8.384' in add_set:
    #     print("'BatteryTypeCode' is in the set")
    # else:
    #     print("'BatteryTypeCode' is not in the set")

    # exit()

    additions_attributes_s7_df = gs1_df_attributes_processed[gs1_df_attributes_processed['FieldID'].isin(add_set)].copy()

    # OneWs attributes
    additions_attributes_onews_s7_df = additions_attributes_s7_df.copy()

    # Setting 'Attribute Type' to 'Common' for all rows
    additions_attributes_onews_s7_df['Attribute Type'] = 'Common'

    # Replacing 'CatSpec' with 'OneWS' in 'Attribute Name' and removing all spaces
    additions_attributes_onews_s7_df['Attribute Name'] = additions_attributes_onews_s7_df['Attribute Name'].str.replace('CatSpec', 'OneWS').str.replace(' ', '')

    # Adjust Is Inheritable for OneWs
    additions_attributes_onews_s7_df['Is Inheritable'] = np.where(
        (additions_attributes_s7_df['Format'] == 'NumberPicklist') | (additions_attributes_s7_df['Format'] == 'Picklist'),
        'YES',  # Set to 'YES' if 'Format' is either 'NumberPicklist' or 'Picklist'
        additions_attributes_onews_s7_df['Is Inheritable']  # Keep the original value if conditions are False
    )

    # Update 'Data Type' and 'Display Type' based on 'Format'
    additions_attributes_s7_df['Data Type'] = np.where(
        additions_attributes_s7_df['Format'] == 'Boolean',
        'Boolean',  # Set 'Data Type' to 'Boolean' if 'Format' is 'Boolean'
        additions_attributes_s7_df['Data Type']  # Otherwise, keep the current value
    )

    additions_attributes_s7_df['Display Type'] = np.where(
        additions_attributes_s7_df['Format'] == 'Boolean',
        'DropDown',  # Set 'Display Type' to 'DropDown' if 'Format' is 'Boolean'
        additions_attributes_s7_df['Display Type']  # Otherwise, keep the current value
    )

    # Adjust Is Inheritable 
    additions_attributes_onews_s7_df['Is Inheritable'] = np.where(
        (additions_attributes_s7_df['Format'] == 'NumberPicklist') | (additions_attributes_s7_df['Format'] == 'Picklist'),
        'YES',  # Set to 'YES' if 'Format' is either 'NumberPicklist' or 'Picklist'
        additions_attributes_onews_s7_df['Is Inheritable']  # Keep the original value if conditions are False
    )

    # Condition to identify rows where 'Is Lookup' is 'YES'
    condition_yes = additions_attributes_onews_s7_df['Is Lookup'] == 'YES'
    condition_not_yes = ~condition_yes  # This is the negation of the condition

    # Apply modifications only to the rows that meet the 'YES' condition
    additions_attributes_onews_s7_df.loc[condition_yes, 'LookUp Table Name'] = additions_attributes_onews_s7_df.loc[condition_yes, 'Attribute Name'].str.replace('_', '').str.slice(0, 30)
    additions_attributes_onews_s7_df.loc[condition_yes, 'LookUp Display Columns'] = '[Code],[Description]'
    additions_attributes_onews_s7_df.loc[condition_yes, 'LookUp Search Columns'] = '[Code],[Description]'
    additions_attributes_onews_s7_df.loc[condition_yes, 'LookUp Display Format'] = '[Code]'
    additions_attributes_onews_s7_df.loc[condition_yes, 'LookUp Sort Order'] = '[Code]'
    additions_attributes_onews_s7_df.loc[condition_yes, 'Export Format'] = '[Code]'

    # Set the mentioned columns to blank '' where 'Is Lookup' is not 'YES'
    additions_attributes_onews_s7_df.loc[condition_not_yes, 'LookUp Table Name'] = ''
    additions_attributes_onews_s7_df.loc[condition_not_yes, 'LookUp Display Columns'] = ''
    additions_attributes_onews_s7_df.loc[condition_not_yes, 'LookUp Search Columns'] = ''
    additions_attributes_onews_s7_df.loc[condition_not_yes, 'LookUp Display Format'] = ''
    additions_attributes_onews_s7_df.loc[condition_not_yes, 'LookUp Sort Order'] = ''
    additions_attributes_onews_s7_df.loc[condition_not_yes, 'Export Format'] = ''

    additions_attributes_onews_s7_df['Attribute Parent Name'] = '' # Logic to be determined, manual for now
    additions_attributes_onews_s7_df['Is Complex'] = '' # Logic to be determined, manual for now
    additions_attributes_onews_s7_df['Is ReadOnly'] = 'YES'

    additions_attributes_onews_s7_df['Is Localizable'] = additions_attributes_onews_s7_df['ONEWS_Is Localizable']
    additions_attributes_onews_s7_df['Is Searchable'] = 'NO'
    additions_attributes_onews_s7_df['Is Null Value Search Required'] = 'NO'
    additions_attributes_onews_s7_df['Minimum Length'] = pd.to_numeric(additions_attributes_onews_s7_df['Min Len'])
    additions_attributes_onews_s7_df['Maximum Length'] = pd.to_numeric(additions_attributes_onews_s7_df['Max Len'])
    additions_attributes_onews_s7_df['UOM Type'] = additions_attributes_onews_s7_df['UOM Type'].replace('Custom UOM', 'GDSN UOM')

    all_additions_attributes_s7_df = pd.concat([all_additions_attributes_s7_df, additions_attributes_s7_df, additions_attributes_onews_s7_df], ignore_index=True)
    
    return all_additions_attributes_s7_df

# Create an empty delete_attributes_s7_df with necessary columns
all_additions_attributes_s7_df = pd.DataFrame(columns=list(gs1_df_attributes_processed.columns))

all_additions_attributes_s7_df = add_attributes_s7(attribute_add_s7_GS1ID_set, all_additions_attributes_s7_df)

s7_add_picklist_ID_set = set(zip(
    all_additions_attributes_s7_df[
        (all_additions_attributes_s7_df['Picklist ID'].notna()) & 
        (all_additions_attributes_s7_df['Picklist ID'] != '') & 
        (all_additions_attributes_s7_df['LookUp Table Name'].notna()) & 
        (all_additions_attributes_s7_df['LookUp Table Name'] != '') & 
        (all_additions_attributes_s7_df['LookUp Table Name'] != 'YesNo') & 
        (all_additions_attributes_s7_df['Display Type'] == 'LookupTable')
    ]['Picklist ID'],
    all_additions_attributes_s7_df[
        (all_additions_attributes_s7_df['Picklist ID'].notna()) & 
        (all_additions_attributes_s7_df['Picklist ID'] != '') & 
        (all_additions_attributes_s7_df['LookUp Table Name'].notna()) & 
        (all_additions_attributes_s7_df['LookUp Table Name'] != '') & 
        (all_additions_attributes_s7_df['LookUp Table Name'] != 'YesNo') &
        (all_additions_attributes_s7_df['Display Type'] == 'LookupTable')
    ]['LookUp Table Name']
))

    ####################
    ## Changes
    ####################
print(f'## Attribute Changes - S7 ##')
overlap_attributes_df = gs1_df_attributes_processed[gs1_df_attributes_processed['FieldID'].isin(attribute_overlap_s7_GS1ID_set)].copy()

# Step 1: Duplicate and prefix selected columns in maxeda_s7_df to compare including the key being attribute code
columns_to_compare = ['Attribute code','Data Type', 'Display Type', 'Precision', 'Is Collection', 'LookUp Table Name'] #'Allowed UOMs'

# Create a new DataFrame, maxeda_s7_changes_df, with only the specified columns from maxeda_s7_df
maxeda_s7_changes_df = maxeda_s7_df_category[columns_to_compare].copy()

# Remove 'Attribute code' to function as the key to join on later
columns_to_compare.remove('Attribute code')

# Prefix check-columns
rename_dict = {col: 'ORIGINAL_' + col for col in columns_to_compare}

# Rename the columns using the dictionary
maxeda_s7_changes_df.rename(columns=rename_dict, inplace=True)

# The original lookup table is needed for determination later on not to compare
columns_to_compare.remove('LookUp Table Name')

# Step 2: Perform a left join from overlap_attributes_df to the modified maxeda_s7_df
merged_df = overlap_attributes_df.merge(
    maxeda_s7_changes_df,
    left_on='FieldID',
    right_on='Attribute code',
    how='left'
)

# Step 3: Compare original and newly added values, collect discrepancies and reasons
discrepancy_details = []
for col in columns_to_compare:
    for index, row in merged_df.iterrows():
        # Trim values and check if they are not empty
        original_trimmed = (str(row['ORIGINAL_' + col]).strip() if pd.notnull(row['ORIGINAL_' + col]) else '')
        new_trimmed = (str(row[col]).strip() if pd.notnull(row[col]) else '')
        
        # Add to discrepancy details only if both are not empty and different
        if (original_trimmed or new_trimmed) and (original_trimmed != new_trimmed):
            discrepancy_details.append({
                'FieldID': row['FieldID'],
                'Column': col,
                'Original Value': row['ORIGINAL_' + col],
                'New Value': row[col],
                'Picklist ID': row['Picklist ID']
            })


# Convert discrepancy details to a DataFrame for better visualization and analysis
discrepancy_df = pd.DataFrame(discrepancy_details)

# Step 4: Extract unique FieldIDs with discrepancies
change_set = set(discrepancy_df['FieldID'].dropna())

# Filter the dataset on the the changed items
changes_s7_df = merged_df[merged_df['FieldID'].isin(change_set)].copy()

# Make set for picklist-value additions
s7_change_to_picklist_ID_set = set(zip(
    changes_s7_df[
        (changes_s7_df['Picklist ID'].notna()) & 
        (changes_s7_df['Picklist ID'] != '') & 
        (changes_s7_df['LookUp Table Name'].notna()) & 
        (changes_s7_df['LookUp Table Name'] != '') & 
        (changes_s7_df['LookUp Table Name'] != 'YesNo') &
        (changes_s7_df['ORIGINAL_Display Type'] != 'LookupTable') & 
        (changes_s7_df['Display Type'] == 'LookupTable')
    ]['Picklist ID'],
    changes_s7_df[
        (changes_s7_df['Picklist ID'].notna()) & 
        (changes_s7_df['Picklist ID'] != '') & 
        (changes_s7_df['LookUp Table Name'].notna()) & 
        (changes_s7_df['LookUp Table Name'] != '') & 
        (changes_s7_df['ORIGINAL_Display Type'] != 'LookupTable') & 
        (changes_s7_df['Display Type'] == 'LookupTable')
    ]['LookUp Table Name']
))

s7_change_from_LookupTableName_set = set(
    changes_s7_df[
        (changes_s7_df['ORIGINAL_LookUp Table Name'] != 'YesNo') & 
        (changes_s7_df['ORIGINAL_Display Type'] == 'LookupTable') & 
        (changes_s7_df['Display Type'] != 'LookupTable')
    ]['ORIGINAL_LookUp Table Name']
)

# print(f"s7_change_to_picklist_set: {s7_change_to_picklist_ID_set}")

# Combine the two DataFrames into one new DataFrame

# Filter columns for the output
columns_s7 = [
    "ID", "Action", "Unique Identifier", "Attribute Type", "Attribute Name",
    "Attribute Long Name", "Attribute Parent Name", "Data Type", "Display Type",
    "Is Collection", "Is Inheritable", "Is Localizable", "Is Complex", "Is Lookup",
    "Is Required", "Is ReadOnly", "Is Hidden", "Show At Entity Creation?", "Is Searchable",
    "Is Null Value Search Required", "Generate Report Table Column?", "Default Value",
    "Minimum Length", "Maximum Length", "Range From", "Is Range From Inclusive",
    "Range To", "Is Range To Inclusive", "Precision", "Use Arbitrary Precision?",
    "UOM Type", "Allowed UOMs", "Default UOM", "Allowable Values", "LookUp Table Name",
    "Lookup Display Columns", "Lookup Search Columns", "Lookup Display Format",
    "Lookup Sort Order", "Export Format", "Sort Order", "Definition", "Example",
    "Business Rule", "Label", "Extension", "Web URI", "Enable History",
    "Apply Time Zone Conversion", "Attribute Regular Expression", "Is UOM Localizable"
]


final_s7_df = pd.concat([delete_attributes_s7_df, all_additions_attributes_s7_df, changes_s7_df], ignore_index=True)
final_s7_additions_and_changes_df = pd.concat([all_additions_attributes_s7_df, changes_s7_df], ignore_index=True)

final_s7_df = final_s7_df[columns_s7]
delete_attributes_s7_df = delete_attributes_s7_df[columns_s7]
final_s7_additions_and_changes_df = final_s7_additions_and_changes_df[columns_s7]
# all_additions_attributes_s7_df = all_additions_attributes_s7_df[columns_s7]
# changes_s7_df = changes_s7_df[columns_s7]

####################
## S8 
####################

print(f'\n  ### Process S8  ###')

    ####################
    ## Delete
    ####################
print(f'## Deletion - S8 ##')

# Initiate the dataset
delete_attributes_s8_df = maxeda_s8_df_scope[maxeda_s8_df_scope['ID'].isin(attribute_delete_s7_MaxedaIDs_set)].copy()
delete_attributes_s8_df['Action'] = 'Delete'

    ####################
    ## Additions
    ####################
print(f'## Additions - S8 ##')

# Initiate the dataset
additions_attributes_s8_df = gs1_df_attributes_processed[gs1_df_attributes_processed['FieldID'].isin(attribute_add_s7_GS1ID_set)].copy()

# Contruct 'Attribute Path'
additions_attributes_s8_df['Attribute Path'] = additions_attributes_s8_df['Attribute Parent Name'] + '//' + additions_attributes_s8_df['Attribute Name']

# Create an empty delete_attributes_s7_df with necessary columns
additons_s8_locale_df = pd.DataFrame(columns=list(additions_attributes_s8_df.columns))

# Create a DataFrame with repeated rows, each original row appearing four times for different locales
locales = ['nl_BE', 'nl_NL', 'fr_BE', 'fr_FR']
additons_s8_locale_df = pd.DataFrame(
    np.repeat(additions_attributes_s8_df.values, len(locales), axis=0),  # Repeat each row
    columns=additions_attributes_s8_df.columns
)

additons_s8_locale_df['Locale'] = np.tile(locales, len(additions_attributes_s8_df))  # Assign locales repeatedly for each group of rows

# Define 'Attribute Long Name' based on 'Locale'
additons_s8_locale_df['Attribute Long Name'] = np.where(
    additons_s8_locale_df['Locale'].isin(['nl_BE', 'nl_NL']),
    additions_attributes_s8_df['Attributename Dutch'].repeat(len(locales)).values,  # Repeat values to match the new row expansion
    additions_attributes_s8_df['Attributename French'].repeat(len(locales)).values
)

#Rename headers
additons_s8_locale_df.rename(columns={'Minimum Length': 'Min Length', 'Maximum Length': 'Max Length'}, inplace=True)

# Empty all values in 'Min Length' and 'Max Length'
additons_s8_locale_df['Min Length'] = ''
additons_s8_locale_df['Max Length'] = ''

# Onews
additions_onews_s8_locale_df = additons_s8_locale_df.copy()
# Replacing 'CatSpec' with 'OneWS' in 'Attribute Name' and removing all spaces
additions_onews_s8_locale_df['Attribute Name'] = additions_onews_s8_locale_df['Attribute Name'].str.replace('CatSpec', 'OneWS').str.replace(' ', '').str.replace('Category Specific Attributes//', 'OneWS_XXXX//')

all_additions_attributes_s8_df = pd.concat([additons_s8_locale_df, additions_onews_s8_locale_df], ignore_index=True)


# Combine all data frames
final_s8_df = pd.concat([delete_attributes_s8_df, all_additions_attributes_s8_df], ignore_index=True)

# Configure output
columns_s8 = ['ID', 'Action', 'Unique Identifier', 'Attribute Path', 'Locale', 'Attribute Long Name', 'Min Length', 'Max Length', 'Definition', 'Example', 'Business Rule']
final_s8_df = final_s8_df[columns_s8]
delete_attributes_s8_df = delete_attributes_s8_df[columns_s8]
all_additions_attributes_s8_df = all_additions_attributes_s8_df[columns_s8]

#######################
## S23 
#######################

print(f'\n  ### Process S23 ###')

# Combine the sets of deleted picklists and exiting attributes changed format FROM Picklist to something else
LookupTable_delete_total_set = attribute_delete_s7_LookupTableName_set.union(s7_change_from_LookupTableName_set)

# Combine the sets of added picklists and exiting attributes changed format TO Picklist from something else
LookupTable_add_total_set = s7_add_picklist_ID_set.union(s7_change_to_picklist_ID_set)

# # Check if 'BatteryTypeCode' is in the set
# if 'BatteryTypeCode' in s7_add_picklist_ID_set:
#     print("'BatteryTypeCode' is in the set")
# else:
#     print("'BatteryTypeCode' is not in the set")

# # Check if 'BatteryTypeCode' is in the set
# if 'BatteryTypeCode' in s7_change_to_picklist_ID_set:
#     print("'BatteryTypeCode' is in the set")
# else:
#     print("'BatteryTypeCode' is not in the set")

# exit()


# Create a new DataFrame with the same columns as maxeda_s23_df
maxeda_s23_total_df = pd.DataFrame(columns=maxeda_s23_df.columns)

    #######################
    ## Delete
    #######################
print(f'## Deletions - S23 ##')

# print(f'LookupTable_delete_total_set {len(LookupTable_delete_total_set)}')

# maxeda_s23_delete_df = maxeda_s23_df[maxeda_s23_df['Table Name'].isin(LookupTable_delete_total_set)].copy()
# Don't delete change from lookuptable as this needs data migration first. Then this script will not pick it up anymore as a change
maxeda_s23_delete_df = maxeda_s23_df[maxeda_s23_df['Table Name'].isin(attribute_delete_s7_LookupTableName_set)].copy()
maxeda_s23_delete_df['Action'] = 'Delete'

maxeda_s23_total_df = pd.concat([maxeda_s23_total_df, maxeda_s23_delete_df], ignore_index=True)


# Create backlog items for S23
backlog_s23_deletions = attributes_aggregation_workflow_lookuptable_df[attributes_aggregation_workflow_lookuptable_df['LookUp Table Name'].isin(attribute_delete_s7_LookupTableName_backlog_set)].copy()
# backlog_s23_deletions = backlog_s23_deletions[backlog_s23_deletions['Category'] != 'YesNo']
backlog_s23_deletions['Level'] = 'Lookup Table (& values)'
backlog_s23_deletions['Reason'] = 'Lookup Table deletion (& its values)'

backlog_s23_deletions['Category'] = (
    backlog_s23_deletions['ID'].astype(str) + ' - ' +
    backlog_s23_deletions['LookUp Table Name']
)

final_backlog_df = pd.concat([final_backlog_df, backlog_s23_deletions], ignore_index=True)

    #######################
    ## Add
    #######################
print(f'## Additions - S23 ##')

# print(LookupTable_add_total_set)
# print(len(LookupTable_add_total_set))

# exit()

all_additions_attributes_s23_df = pd.DataFrame(columns=list(maxeda_s23_delete_df.columns))

for picklist_id, lookup_table_name in LookupTable_add_total_set:
    
    # Determine the common row values
    table_name = lookup_table_name
    sequence = 0
    column_name_first_row = "Code"
    column_name_second_row = lookup_table_name

    # Find the corresponding data type and precision from gs1_df_attributes_processed
    match = gs1_df_attributes_processed[gs1_df_attributes_processed['Picklist ID'] == picklist_id]
    data_type = match['Data Type'].values[0] if not match.empty else None

    # # Check if 'Attributename English' column has the value 'Battery size type code'
    # mask = match['Attributename English'] == 'Battery size type code'

    # # Print 'Picklist ID' for rows where the condition is true
    # if mask.any():
    #     print(match[mask]['Picklist ID'].values[0])
    #     exit()
    
    width = 500
    precision = 0
    nullable_first_row = "Yes"
    nullable_second_row = "Yes"
    is_unique_first_row = "Yes"
    is_unique_second_row = "No"
    
    if lookup_table_name.startswith("OneWS"):
        column_name_second_row = "Description"
        width = 250
        nullable_first_row = "No"

    # Create the first row
    first_row = {
        'Table Name': table_name,
        'Sequence': sequence,
        'Column Name': column_name_first_row,
        'Data Type': data_type,
        'Width': width,
        'Precision': precision,
        'Nullable?': nullable_first_row,
        'Is Unique': is_unique_first_row
    }
    
    # Create the second row
    second_row = {
        'Table Name': table_name,
        'Sequence': sequence,
        'Column Name': column_name_second_row,
        'Data Type': data_type,
        'Width': width,
        'Precision': precision,
        'Nullable?': "Yes",
        'Is Unique': is_unique_second_row
    }
    
    # Append the rows to the new DataFrame
    # maxeda_s23_total_df = maxeda_s23_total_df.append(first_row, ignore_index=True)
    # maxeda_s23_total_df = maxeda_s23_total_df.append(second_row, ignore_index=True)
    
    all_additions_attributes_s23_df = pd.concat([all_additions_attributes_s23_df, pd.DataFrame([first_row]), pd.DataFrame([second_row])], ignore_index=True)

    maxeda_s23_total_df = pd.concat([maxeda_s23_total_df, all_additions_attributes_s23_df], ignore_index=True)

# # Check if 'Attributename English' column has the value 'Battery size type code'
# mask = all_additions_attributes_s23_df['Attributename English'] == 'Battery size type code'

# # Print 'Picklist ID' for rows where the condition is true
# if mask.any():
#     print(all_additions_attributes_s23_df[mask]['Picklist ID'].values[0])
#     exit()

#######################
## S14
#######################

print(f'\n  ### Process S14 ###')

# Obtain the GS1 attribute ID
maxeda_s14_df = pd.merge(
    maxeda_s14_df,
    maxeda_s8_full_df[['Attribute Path', 'ID']],
    how='left',
    on = 'Attribute Path'
)

maxeda_s14_df = maxeda_s14_df.drop_duplicates()   

maxeda_s14_df = pd.merge(
    maxeda_s14_df,
    maxeda_s7_df_scope[['ID', 'Attribute code']],
    how='left',
    left_on = 'ID_y',
    right_on = 'ID'
)

# Remove rows where 'Attribute code' is NaN
maxeda_s14_df = maxeda_s14_df.dropna(subset=['Attribute code'])
# Exclude Maxeda bricks
maxeda_s14_df = maxeda_s14_df[~maxeda_s14_df['Category Name'].str.startswith('999')]
# Exclude maxeda-attributes
maxeda_s7_df_scope = maxeda_s7_df_scope[~maxeda_s7_df_scope['Attribute code'].str.startswith("M")]


# Create a sets to compare
maxeda_brick_attribute_set = set(zip(maxeda_s14_df['Category Name'], maxeda_s14_df['Attribute code']))
gs1_brick_attribute_set = set(zip(gs1_df_attributes_brick_active['Brick'], gs1_df_attributes_brick_active['FieldID']))

# # Define the target substring
# target_substring = '10001688'

# # Check if the target substring is in the first item of any tuple in maxeda_brick_attribute_set
# contains_target_substring = any(target_substring in item[0] for item in maxeda_brick_attribute_set)
# print(f"Does any 'Category Name' in maxeda_brick_attribute_set contain '{target_substring}'? {contains_target_substring}")
# contains_target_substring = any(target_substring in item[0] for item in gs1_brick_attribute_set)
# print(f"Does any 'Category Name' in gs1_brick_attribute_set contain '{target_substring}'? {contains_target_substring}")

brick_attribute_delete_temp_set = maxeda_brick_attribute_set - gs1_brick_attribute_set
brick_attribute_add_set = gs1_brick_attribute_set - maxeda_brick_attribute_set
# brick_attribute_overlap_set = gs1_brick_attribute_set & maxeda_brick_attribute_set

# contains_target_substring = any(target_substring in item[0] for item in brick_attribute_delete_temp_set)
# print(f"Does any 'Category Name' in brick_attribute_delete_temp_set contain '{target_substring}'? {contains_target_substring}")
# exit()

# print("maxeda_brick_attribute_set (first 10 items):", list(maxeda_brick_attribute_set)[:10])
# print("gs1_brick_attribute_set (first 10 items):", list(gs1_brick_attribute_set)[:10])
# print(f'brick_attribute_delete_temp_set: {list(brick_attribute_delete_temp_set)[:10]}')
# print(f'brick_attribute_add_set: {list(brick_attribute_add_set)[:10]}')
# print(f'brick_attribute_overlap_set: {list(brick_attribute_overlap_set)[:10]}')

# Remove the workflow brick-attribute combinations from the delete set. It will not be put into the backlog as Brick and/or attribute deletion already is.
# Only put those items on the backlog that do not belong to brick deletion or attribute deletion. 
    # These are allready on the backlog and will be organically processed
    # This way only the removed combinations will be present in the table. I.o.w. the attribute still exists but in other bricks.

# Add the attribute to brick_aggregation_workflow_with_attributes_df which currently only contains the maxeda attribute ID
# brick_aggregation_workflow_with_GS1attributes_df = pd.merge(
#     brick_aggregation_workflow_with_attributes_df, 
#     maxeda_s7_df_scope[['ID', 'Attribute code']], 
#     how='left', 
#     on='ID',              
# )

# print(brick_aggregation_workflow_with_GS1attributes_df)
# exit()

# Produce set to exclude from deletion
# print(Attribute_allIDs_aggregation_workflow_df.head())
brick_attribute_delete_workflow_set = set(zip(brick_attributes_aggregation_workflow_df['Brick'], brick_attributes_aggregation_workflow_df['Attribute code']))
brick_attribute_delete_set = brick_attribute_delete_temp_set - brick_attribute_delete_workflow_set
brick_attribute_delete_backlog_set = brick_attribute_delete_temp_set - brick_attribute_delete_set

# print(f'brick_attribute_delete_workflow_set: {list(brick_attribute_delete_workflow_set)[:10]}')
# print(f'brick_attribute_delete_set: {list(brick_attribute_delete_set)[:10]}')
# print(f'brick_attribute_delete_backlog_set: {list(brick_attribute_delete_backlog_set)[:10]}')


    #######################
    ## Delete
    #######################
print(f'## Deletions - S14 ##')
# Use a boolean mask to filter S14 rows for deletion
mask = maxeda_s14_df.apply(
    lambda row: (row['Category Name'], row['Attribute code']) in brick_attribute_delete_set,
    axis=1
)

maxeda_s14_delete_df = maxeda_s14_df[mask].copy()
maxeda_s14_delete_df['Action'] = 'Delete'

## Create backlog
# To be deleted bricks and attributes are already included in their respective backlogs and should therefore be excluded in this backlog
    # A next run of the script, after resovling the backlogs of brick and attributes, will organically delete the 'old' combinations

# Create a set from the brick backlog.
bricks_to_exclude_set = set(final_backlog_df.loc[final_backlog_df['Reason'] == 'Brick deletion', 'Category'])

# Filter the set to exclude tuples where the first item is in categories_to_exclude
final_brick_attribute_delete_backlog_temp_set = {
    item for item in brick_attribute_delete_backlog_set
    if item[0] not in bricks_to_exclude_set
}

# Create a set from the attribute backlog
attributes_to_exclude_df = backlog_attribute_deletions.merge(
    maxeda_s7_df_scope[['ID', 'Attribute code']],
    on='ID',
    how='left'
)         

# print(attributes_to_exclude_df.head())

attributes_to_exclude_set = set(attributes_to_exclude_df['Attribute code_x'])

# Filter the set to exclude tuples where the first item is in categories_to_exclude
final_brick_attribute_delete_backlog_set = {
    item for item in final_brick_attribute_delete_backlog_temp_set
    if item[1] not in attributes_to_exclude_set
}

# print(len(brick_attribute_delete_backlog_set))
# print(len(final_brick_attribute_delete_backlog_set))
# exit()

# print(f'maxeda_s14_df: {len(maxeda_s14_df)}')
# print(f'maxeda_s14_delete_df: {len(maxeda_s14_delete_df)}')
# print(f'bricks_to_exclude: {len(bricks_to_exclude_set)}')
# print(f'attributes_to_exclude: {len(attributes_to_exclude_set)}')

# print(f'maxeda_s14_delete_df: {maxeda_s14_delete_df[:10]}')
# print(f'brick_attribute_delete_backlog_set: {list(brick_attribute_delete_backlog_set)[:10]}')
# print(f'bricks_to_exclude: {bricks_to_exclude_set}')
# print(f'attributes_to_exclude: {attributes_to_exclude_set}')
# print(f'final_brick_attribute_delete_backlog_set: {list(final_brick_attribute_delete_backlog_set)[:10]}')


# print(attributes_aggregation_workflow_GS1_ID_df.columns)
# Use a boolean mask to filter the rows
mask = brick_attributes_aggregation_workflow_df.apply(
    lambda row: (row['Brick'], row['Attribute code']) in final_brick_attribute_delete_backlog_set,
    axis=1
)

# Select the rows based on the mask
brick_attribute_backlog = brick_attributes_aggregation_workflow_df[mask].copy()
brick_attribute_backlog['Level'] = 'Brick-Attribute'
brick_attribute_backlog['Reason'] = 'Delete Brick-Attribute'

# Create the 'Category' column by concatenating 'Brick' and 'Attribute Path'
brick_attribute_backlog['Category'] = brick_attribute_backlog['Brick'] + ' - ' + brick_attribute_backlog['Attribute code'] + ' - ' + brick_attribute_backlog['Attribute Path']

# Select and reorder the columns
selected_columns = ['Category', 'unique_count_GTIN', 'unique_count_VendorNumberSAP', 'unique_count_ArticleLongName', 'Level', 'Reason']
brick_attribute_backlog = brick_attribute_backlog[selected_columns]

final_backlog_df = pd.concat([final_backlog_df, brick_attribute_backlog], ignore_index=True)

# print("Filtered brick_aggregation_workflow_with_GS1attributes_df:")
# print(brick_attribute_backlog.columns)
# # exit()

# # print(len(Attribute_allIDs_aggregation_workflow_df))
# print(len(brick_attribute_backlog))
# print(brick_attribute_backlog.head)

# Define the path to save the Excel file
output_file_path = 'filtered_brick_aggregation.xlsx'

# Write the DataFrame to an Excel file
brick_attribute_backlog.to_excel(output_file_path, index=False)

# exit()

# print("maxeda_s14_delete_df:")
# print(maxeda_s14_delete_df)

# print("brick_aggregation_workflow_with_GS1attributes_df:")
# print(Attribute_allIDs_aggregation_workflow_df[:10])

# Only put those items on the backlog that do not belong to brick deletion or attribute deletion. 
    # These are allready on the backlog and will be organically processed
    # This way only the removed combinations will be present in the table. I.o.w. the attribute still exists but in other bricks.

# exit()


# add the attribute ID based on 'Attribute Path' --> S8 original 'Attribute Path' --> S8 'ID' --> S7 enriched with attribute ID 'ID' --> S7 enriched with attribute ID 'Attribute ID'
# Product as-is set
# Look for or produce GS1 brick-attribute st

# Compare the sets
# Deletions incl. backlog based on bricks. For backlog exclude attributes that were deleted entirely. I.o.w. only include bricks that have excluded an attribute that continues to exist
# Additions
# Changes of Brick hierarchie 

    #######################
    ## Add
    #######################
print(f'## Additions - S14 ##')
# Initialize an empty list to store the rows for the new DataFrame
new_rows = []

# print(new_brick_categories_s9.columns)
# print(maxeda_s9.columns)
# # exit()


### Concatenate the DataFrames old + new to find the corresponding meta-data
##Bricks
# Select only the necessary columns from both dataframes
new_brick_categories_selected = new_brick_categories_s9[['Category Name', 'Parent Category Path']]
maxeda_bricks_selected = maxeda_s9[['Category Name', 'Parent Category Path']]
# Add changes because the hierarchie could have changed
change_brick_categories_selected = change_hierarchy_bricks_s9[['Category Name', 'Parent Category Path']]
# Concatenate the selected columns
parent_category_path_lookup_df = pd.concat([new_brick_categories_selected, change_brick_categories_selected, maxeda_bricks_selected], ignore_index=True)
# Drop duplicates based on 'Category Name', keeping the first occurrence. This takes into account that first hierarchy changs of bricks have been processed manually
unique_parent_category_path_lookup_df = parent_category_path_lookup_df.drop_duplicates(subset=['Category Name'], keep='first')
# Reset the index
unique_parent_category_path_lookup_df.reset_index(drop=True, inplace=True)

## Attributes
# Select only the necessary columns from both dataframes
# print(all_additions_attributes_s7_df.head())
# print(maxeda_s7_df_scope.head())
new_attributes_categories_selected = all_additions_attributes_s7_df[['FieldID', 'Attribute Name', 'Attribute Parent Name', 'Is Required', 'Allowed UOMs', 'Default UOM']]
new_attributes_categories_selected = new_attributes_categories_selected.rename(columns={'FieldID': 'Attribute code'})
maxeda_attributes_selected = maxeda_s7_df_scope[['Attribute code', 'Attribute Name', 'Attribute Parent Name', 'Is Required', 'Allowed UOMs', 'Default UOM']]
# Concatenate the selected columns
attribute_path_lookup_df = pd.concat([new_attributes_categories_selected, maxeda_attributes_selected], ignore_index=True)
# Only keep the Catspe_attributes
attribute_path_lookup_df= attribute_path_lookup_df[attribute_path_lookup_df['Attribute Name'].str.startswith('CatSpec_')]


# Loop through the set and filter the DataFrame
for brick, field_id in brick_attribute_add_set:
    # print(brick, field_id)
    attribute_metadata = attribute_path_lookup_df[attribute_path_lookup_df['Attribute code'] == field_id]
    filtered_new_brick_row = unique_parent_category_path_lookup_df[unique_parent_category_path_lookup_df['Category Name'] == brick]
    
    if not attribute_metadata.empty:
        # Extract the required columns and append to the new_rows list
        new_row = {
            'ID': '',
            'Action': '',
            'Unique Identifier': '',
            'Hierarchy Name': '',
            'Category Name': brick,
            'Parent Category Path': filtered_new_brick_row['Parent Category Path'].values[0],
            'Attribute Path': attribute_metadata['Attribute Parent Name'].values[0] + '//' + attribute_metadata['Attribute Name'].values[0],
            'Attribute Long Name': '',
            'Is Required': attribute_metadata['Is Required'].values[0],
            'Is ReadOnly': '',
            'Default Value': '',	
            'Minimum Length': '',	
            'Maximum Length': '',	
            'Range From': '',	
            'Is Range From Inclusive': '',	
            'Range To': '',
            'Is Range To Inclusive': '',	
            'Precision': '',
            'Allowed UOMs': attribute_metadata['Allowed UOMs'].values[0],
            'Default UOM': attribute_metadata['Default UOM'].values[0],
            'Allowable Values': '',	
            'Sort Order': '',	
            'Definition': '',	
            'Example': '',	
            'Business Rule': '',
            'Inheritable Only': 'NO',	
            'Auto Promotable': 'NO'
        }

        new_rows.append(new_row)

# Create a new DataFrame from the list of rows
add_S14_df = pd.DataFrame(new_rows)

# Display the new DataFrame
# print(add_S14_df)

# exit()



####################
## LookupData values
####################
print(f'\n  ### Process - LookupData values ###')
# print(f'### Lookup tables ###')
maxeda_lookuptables_s7_set = set(maxeda_s7_df_scope['LookUp Table Name'].replace('', np.nan).dropna())

valid_combinations = set(zip(gs1_df_picklists['Picklist ID'], gs1_df_picklists['Code value']))

def read_excel_with_openpyxl(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook[sheet_name]
    data = []

    for row in sheet.iter_rows(values_only=True):
        data.append([str(cell).strip() if cell is not None else "" for cell in row])

    # Check if the required headers are present and if the second column has at least one non-empty value
    if data and len(data[0]) >= 2 and data[0][0] == "Id" and "Code" in data[0][1]:
        df = pd.DataFrame(data[1:], columns=data[0])
        return df
    return None  # Return None if conditions are not met

def read_all_excel_files(directory_path):
    counter = 0 # for testing purposes

    final_delete_lookup_values_df = pd.DataFrame()
    final_delete_lookup_values_backlog_df = pd.DataFrame()
    final_add_lookup_values_df = []
    
    # Get a list of .xlsx files in the directory
    xlsx_files = [filename for filename in os.listdir(directory_path) if filename.endswith('.xlsx')]
    
    for filename in tqdm(xlsx_files, desc="  ### Process Lookup table value files ###"):
        counter += 1 # for testing purposes
        if counter < 1000: # for testing purposes
            # print(f"counter LookupData files: {counter}")
            file_path = os.path.join(directory_path, filename)
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            # Iterate over each sheet except the first one
            for sheet_name in workbook.sheetnames[1:]:
                if sheet_name in maxeda_lookuptables_s7_set: # only consider the sheets that are LookupTables for s7-categoeie variables

                    disregarded_entries = []
                    added_entries = []

                    df = read_excel_with_openpyxl(file_path, sheet_name) # read the Excel-sheet
                    if df is not None:
                        # Determine GS1 field ID
                        attribute_codes = maxeda_s7_df_scope.loc[maxeda_s7_df_scope['LookUp Table Name'] == sheet_name, 'Attribute code'] # Lookup the GS1-attribute code for this LookupData
                        attribute_code = attribute_codes.iloc[0] 

                        if attribute_code == '':
                            continue                       

                        # Next, determine the GS1 picklist id
                        # print("# Lookup picklist id #")
                        picklist_ids = gs1_df_attributes.loc[gs1_df_attributes['FieldID'] == attribute_code, 'Picklist ID']
                        if picklist_ids.empty:
                            # print(f"No Picklist ID found for attribute code: {attribute_code}. Skipping to next sheet.")
                            picklist_id = None
                        else:
                            picklist_id = picklist_ids.iloc[0]
    
                        # Determine the valid values for this picklist
                        filtered_valid_combinations = {code for pid, code in valid_combinations if pid == picklist_id}

                        # print("# Delete rows #")
                    
                        
                        ##############################
                        ## Delete
                        ##############################
                        # print(f'## Deletions - Lookuptable values ##')
                        # Define a function to filter rows and collect details of disregarded entries
                        def delete_combinations(row, filtered_valid_combinations):
                            # Check if the combination of 'Picklist ID' and the value in the second column is in the valid combinations
                            if (row.iloc[1]) not in filtered_valid_combinations:
                                # Collect filename, sheet name, and the second column's value for disregarded entries
                                disregarded_entries.append({
                                    'filename': filename,  # Assumes 'filename' is defined where this function is called
                                    'sheet_name': sheet_name,  # Assumes 'sheet_name' is defined where this function is called
                                    'value': row.iloc[1]
                                })
                                return False
                            return True

                        df.apply(delete_combinations, axis=1, args=(filtered_valid_combinations,))


                            ##############################
                            ## Integrate backlog
                            ##############################                       

                        # Create a set of unique values from the 'LookUp Table Name' column
                        lookup_table_name_workflow_set = set(attributes_aggregation_workflow_lookuptable_df['LookUp Table Name'].dropna())

                        ## Final deletions
                        # filter the backlog items out of the disregarded_entries
                        filtered_disregarded_entries = [
                            entry for entry in disregarded_entries if entry['sheet_name'] not in lookup_table_name_workflow_set
                        ]

                        # if picklist_id == '6.660':
                        #     print(f'filtered_valid_combinations: {filtered_valid_combinations}')
                        #     print(f'disregarded_entries: {disregarded_entries}')
                        #     print(f'filtered_disregarded_entries: {filtered_disregarded_entries}')
                        #     exit()    

                        # Convert the filtered list to a DataFrame
                        disregarded_entries_df = pd.DataFrame(filtered_disregarded_entries)


                        # print(disregarded_entries_df)
                        # Restructure the table
                        if not disregarded_entries_df.empty:
                            disregarded_entries_df.drop(columns=['filename'], inplace=True)
                            disregarded_entries_df.rename(columns={'sheet_name': 'LookUp Table'}, inplace=True)

                            # Concatenate the current disregarded_entries_df with final_disregarded_entries_df
                            final_delete_lookup_values_df = pd.concat([final_delete_lookup_values_df, disregarded_entries_df], ignore_index=True)
                        

                        ## Backlog
                        # Filter the disregarded entries based on the set
                        lookupvalues_backlog_list = [
                            entry for entry in disregarded_entries 
                            if entry['sheet_name'] in lookup_table_name_workflow_set 
                            and entry['sheet_name'] not in attribute_delete_s7_LookupTableName_backlog_set # Because it was already added to that backlog
                        ]

                        
                        # Transform the filtered entries into the desired DataFrame format
                        lookupvalues_backlog_df = pd.DataFrame(lookupvalues_backlog_list)  # Define this within your loop

                        # if not lookupvalues_backlog_df.empty:
                        #     print(lookupvalues_backlog_df)
                        #     # exit()

                        # Add the required columns
                        lookupvalues_backlog_df.insert(0, 'Category', sheet_name)  # Insert 'Category' at the first position
                        lookupvalues_backlog_df['Level'] = 'Lookup value'
                        lookupvalues_backlog_df['Reason'] = 'Lookup value deletion(s)'                  

                        if sheet_name == 'Sowingperiod':
                            print(len(lookupvalues_backlog_df))
                        # Merge with Attribute_allIDs_aggregation_workflow_df
                        lookupvalues_backlog_df = lookupvalues_backlog_df.merge(
                            attributes_aggregation_workflow_lookuptable_df,
                            left_on='Category',
                            right_on='LookUp Table Name',
                            how='left'
                        )            
                        if sheet_name == 'Sowingperiod':
                            print(len(lookupvalues_backlog_df))
                            # exit()
                        # # if counter == 90:
                        # print(lookupvalues_backlog_df.columns)
                        # print(lookupvalues_backlog_df.head())

                        # # Example function to check for blank 'value' field in a DataFrame
                        # def check_blank_value_in_dataframe(df):
                        #     has_blank_value = False
                        #     for index, row in df.iterrows():
                        #         if pd.isnull(row['value']) or row['value'] == '':
                        #             has_blank_value = True
                        #             break
                        #     return has_blank_value

                        # # Example usage:
                        # # Assuming disregarded_entries is populated somewhere in your code
                        # is_value_blank = check_blank_value_in_dataframe(lookupvalues_backlog_df)

                        # if is_value_blank:
                        #     print("At least one item in disregarded_entries has a blank value. {}")
                        # else:
                        #     print("No item in disregarded_entries has a blank value.")
                        # # exit()

                        if not lookupvalues_backlog_df.empty:
                            lookupvalues_backlog_df['Category'] = lookupvalues_backlog_df['Category'] + ' - ' + lookupvalues_backlog_df['value']
                        # except KeyError as e:
                        #     print(f"Error accessing 'value' column: {e}")
                        
                        # lookupvalues_backlog_df['Category'] = lookupvalues_backlog_df['Category'] + ' - ' + lookupvalues_backlog_df['value'].astype(str)
                          
                        # Select and reorder the specified columns
                        required_columns = ['Category', 'unique_count_GTIN', 'unique_count_VendorNumberSAP', 'unique_count_ArticleLongName', 'Level', 'Reason']
                        lookupvalues_backlog_df = lookupvalues_backlog_df[required_columns]

                        # if not lookupvalues_backlog_df.empty:
                        #     print(lookupvalues_backlog_df)
                        #     # exit()

                        # De-duplicate
                        lookupvalues_backlog_df = lookupvalues_backlog_df.drop_duplicates()   

                        # Add to full backlog
                        final_delete_lookup_values_backlog_df = pd.concat([final_delete_lookup_values_backlog_df, lookupvalues_backlog_df], ignore_index=True)

                        # if not lookupvalues_backlog_df.empty:
                        #     print(final_delete_lookup_values_backlog_df)
                        #     # exit()
                        
                        

                        ##############################
                        ## Add
                        ##############################
                        # print(f'## Additions - Lookuptable data ##')
                        # # Check if the DataFrame has at least two columns
                        if len(df.columns) < 2:     
                            print(f"Not enough columns in {filename} - {sheet_name}.")
                            continue
                        
                        # Check if the DataFrame is empty
                        if not df.empty:
                            current_combinations = set(df.iloc[:, 1])
                        else:
                            current_combinations = set()
                            # print(f"No data in {filename} - {sheet_name}.")

                        # Determine missing combinations
                        missing_combinations = filtered_valid_combinations - current_combinations

                        # Define the language codes and their corresponding columns in gs1_df_picklists
                        language_mapping = {
                            "en_US": "Values in English used for user interface ",
                            "nl_NL": "Values in Dutch used for user interface ",
                            "nl_BE": "Values in Dutch used for user interface ",
                            "fr_FR": "Values in French used for user interface ",
                            "fr_BE": "Values in French used for user interface "
                        }
                
                        for code_value in missing_combinations: 
                            added_entries.append({
                                    'filename': filename,  # Assumes 'filename' is defined where this function is called
                                    'sheet_name': sheet_name,  # Assumes 'sheet_name' is defined where this function is called
                                    'value': code_value
                                })
                            
                            # Find the row in gs1_df_picklists corresponding to the current code_value and picklist_id
                            row_data = gs1_df_picklists[(gs1_df_picklists['Picklist ID'] == picklist_id) & (gs1_df_picklists['Code value'] == code_value)]

                            # Initialize the new row with the code value
                            new_row = {df.columns[1]: code_value}  # filtered_df.columns[1] should be the column where code values are stored

                            # Assign language specific values
                            for language_code, gs1_col in language_mapping.items():
                                # Find columns in df that contain the language code
                                matched_columns = [col for col in df.columns if language_code in col]
                                # Get the value from gs1_df_picklists
                                language_value = row_data[gs1_col].values[0] if not row_data.empty else ''
                                # Assign this value to all matched columns in new_row
                                for matched_col in matched_columns:
                                    new_row[matched_col] = language_value

                            # Append the new row to df
                            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

                            # # Check if new_row has a value
                            # if new_row[df.columns[1]]:
                            #     print(new_row)
                            #     print(df)
                            #     exit()

                        # Make a total dataframe
                        if added_entries:
                            final_add_lookup_values_df.append({'df': df, 'filename': filename, 'sheet_name': sheet_name, 'Picklist': picklist_id})

    return final_delete_lookup_values_df, final_delete_lookup_values_backlog_df, final_add_lookup_values_df  # Return the final DataFrames

directory_path = os.getenv('lookup_data')
final_delete_lookup_values, final_delete_lookup_values_backlog, final_add_lookup_values = read_all_excel_files(directory_path)
# delete_lookupdata_df = final_lookupdata_df.copy()

# print(final_add_lookup_values)
# exit()

final_backlog_df = pd.concat([final_backlog_df, final_delete_lookup_values_backlog], ignore_index=True)

# print(final_delete_lookup_values)
# print(final_delete_lookup_values_backlog)
# exit()


#######################
## CDB - MDM Mapping 
#######################
print(f'\n  ### Process - CDB MDM Mapping ###')

    ##################
    ## Add
    ##################
print(f'## Additions - CDB MDM Mapping ##')
# Calculate the midpoint to split the DataFrame into two halves
midpoint = len(all_additions_attributes_s7_df) // 2

# Split the DataFrame into two halves to divide Catspec from OneWS
first_half = all_additions_attributes_s7_df.iloc[:midpoint].reset_index(drop=True)
second_half = all_additions_attributes_s7_df.iloc[midpoint:].reset_index(drop=True)

# Attach 'Attribute Name' of the second half to the first half
first_half['Attribute Name OneWS'] = second_half['Attribute Name']
first_half['Is Localizable OneWS'] = second_half['Is Localizable']

# Create the final DataFrame with specified headers and values
cdb_mdm_add_df = pd.DataFrame({
    'Id': '',
    'MDMAttributeShortName *': first_half['Attribute Name'],
    'MDMAttributeParentName//en_US': first_half['Attribute Parent Name'],
    'GS1AttributeShortName//en_US': first_half['Attribute Name OneWS'],
    'ContentType//en_US': 'Value',
    'NeedUOMConversion//en_US': '', # to be checked
    'TargetUOM//en_US': first_half['Default UOM'],
    'IsLocalized//en_US': first_half['Is Localizable OneWS'],
    'PackageType//en_US': 'BASE_UNIT_OR_EACH',
    'IsChild//en_US': ''  # to be checked
})

# print(cdb_mdm_add_df)

# print(second_half.columns)
# exit()

    ##################
    ## Delete
    ##################

print(f'## Deletions - CDB MDM Mapping ##')
# # Filter the set to keep only strings that start with "CatSpec_"
# attribute_delete_s7_AttributeName_cdb_mdm_set = {s for s in attribute_delete_s7_AttributeName_set if s.startswith("CatSpec_")}

cdb_mdm_delete_df = CDB_MDM_mapping_df[CDB_MDM_mapping_df['MDMAttributeShortName *'].isin(attribute_delete_s7_AttributeName_set)].copy()

    ##################
    ## Backlog
    ##################
# attribute_backlog_s7_AttributeName_cdb_mdm_set = {s for s in attribute_delete_s7_AttributeName_backlog_set if s.startswith("CatSpec_")}

cdb_mdm_backlog_df = attributes_aggregation_workflow_ALL_ID_df[attributes_aggregation_workflow_ALL_ID_df['Attribute Name'].isin(attribute_delete_s7_AttributeName_backlog_set)].copy()

cdb_mdm_backlog_df['Category'] = (
    cdb_mdm_backlog_df['Attribute code'] + ' - ' +
    cdb_mdm_backlog_df['ID'].astype(str) + ' - ' +
    cdb_mdm_backlog_df['Attribute Name'] + ' - ' +
    cdb_mdm_backlog_df['Attribute Name'].str.replace('CatSpec', 'OneWS').str.replace(' ', '')
)

cdb_mdm_backlog_df['Level'] = 'Mapping CDB-MDM-GS1'
cdb_mdm_backlog_df['Reason'] = 'Deletion Mapping CDB-MDM-GS1'

# print(cdb_mdm_delete_df)
# print(cdb_mdm_backlog_df)
# print(cdb_mdm_backlog_df.columns)

final_backlog_df = pd.concat([final_backlog_df, cdb_mdm_backlog_df], ignore_index=True)


# exit()


#######################
## CDB - GS1 Mapping (GS1 to maxeda attribute mapping)
#######################
print(f'\n  ### Process - CDB GS1 Mapping ###')
 
    ##################
    ## Add
    ##################
print(f'## Additions - CDB GS1 Mapping ##')
# # Calculate the midpoint to split the DataFrame into two halves
# midpoint = len(all_additions_attributes_s7_df) // 2

# # Split the DataFrame into two halves to divide Catspec from OneWS
# first_half = all_additions_attributes_s7_df.iloc[:midpoint].reset_index(drop=True)
# second_half = all_additions_attributes_s7_df.iloc[midpoint:].reset_index(drop=True)

# # Attach 'Attribute Name' of the second half to the first half
# first_half['Attribute Name OneWS'] = second_half['Attribute Name']
# first_half['Is Localizable OneWS'] = second_half['Is Localizable']
# print(second_half)
# exit()

# Create the final DataFrame with specified headers and values
cdb_gs1_add_df = pd.DataFrame({
    'Id': '',
    'LinkId *': '',
    'GS1Tag//en_US': second_half['TagName'],
    'GS1AttributeShortName//en_US': second_half['Attribute Name'],
    'GS1AttributeParentName//en_US': second_half['Attribute Parent Name'],
    'FieldId//en_US': second_half['FieldID'],
    'IsChild//en_US': ''  # to be checked
})

# print(cdb_gs1_add_df)



    ##################
    ## Delete
    ##################

# Filter the set to keep only strings that start with "CatSpec_"
# attribute_delete_s7_AttributeName_cdb_gs1_set = {s for s in attribute_delete_s7_AttributeName_set if s.startswith("OneWS_")}

# print(f'attribute_delete_s7_AttributeName_set: {attribute_delete_s7_AttributeName_set}\n')
# print(f'attribute_delete_s7_AttributeName_cdb_gs1_set: {attribute_delete_s7_AttributeName_cdb_gs1_set} \n')
# print(f'CDB_GS1_mapping_df: {CDB_GS1_mapping_df}\n')
# print(CDB_GS1_mapping_df['GS1AttributeShortName//en_US'])

cdb_gs1_delete_df = CDB_GS1_mapping_df[CDB_GS1_mapping_df['GS1AttributeShortName//en_US'].isin(attribute_delete_s7_AttributeName_set)].copy()

# print(f'cdb_gs1_delete_df: {cdb_gs1_delete_df}')
#     ##################
#     ## Backlog
#     ##################
# attribute_backlog_s7_AttributeName_cdb_gs1_set = {s for s in attribute_delete_s7_AttributeName_backlog_set if s.startswith("OneWS_")}

# cdb_gs1_backlog_df = attributes_aggregation_workflow_ALL_ID_df[attributes_aggregation_workflow_ALL_ID_df['Attribute Name'].isin(attribute_backlog_s7_AttributeName_cdb_gs1_set)].copy()

# cdb_gs1_backlog_df['Category'] = (
#     cdb_gs1_backlog_df['Attribute code'] + ' - ' +
#     cdb_gs1_backlog_df['ID'].astype(str) + ' - ' +
#     cdb_gs1_backlog_df['Attribute Name']
# )

# print(cdb_gs1_delete_df)
# print(cdb_gs1_backlog_df)
# print(cdb_gs1_backlog_df.columns)

# final_backlog_df = pd.concat([final_backlog_df, cdb_gs1_backlog_df], ignore_index=True)


#######################
## New lookup values for new lookup tables
#######################

# Loop over the set of tuples
for picklist_id, lookup_table_name in LookupTable_add_total_set:

    # Filter gs1_df_picklists for rows with the current Picklist ID
    new_picklist_values_df = gs1_df_picklists[gs1_df_picklists['Picklist ID'] == picklist_id].copy()

    # Create a new DataFrame with the required columns and headers
    reconfigured_df = pd.DataFrame({
        'Id': [''] * len(new_picklist_values_df),
        'Code': new_picklist_values_df['Code value'],
        f'{lookup_table_name}//en_US': new_picklist_values_df['Values in English used for user interface '],
        f'{lookup_table_name}//nl_BE': new_picklist_values_df['Values in Dutch used for user interface '],
        f'{lookup_table_name}//nl_NL': new_picklist_values_df['Values in Dutch used for user interface '],
        f'{lookup_table_name}//fr_BE': new_picklist_values_df['Values in French used for user interface '],
        f'{lookup_table_name}//fr_FR': new_picklist_values_df['Values in French used for user interface ']
    })

#     # print(f"lookup table name: {lookup_table_name}")

    # Add to final dataframe 
    final_add_lookup_values.append({'df': reconfigured_df, 'filename': 'NO FILE: addition', 'sheet_name': lookup_table_name, 'Picklist': picklist_id})

###################
## Write output total
####################
print('\n  ### Output Attributes ###')

print('## Write - S9 & S10 total ##')
# Write the concatenated DataFrame to an Excel file
output_file_path = os.path.join(output_folder, 'GS1_vs_Datamodel_Comparison_Bricks.xlsx')

metadata_df = pd.DataFrame({
    'Sheet No': ['S9', 'S10'],
    'DataModel Type Name': ['Category', 'Category - Localized'],
    'Physical Sheet Name': ['S9 - Category', 'S10 - Category Locale'],
    'Load Lookup?': ['NO', 'NO']
})

with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
    metadata_df.to_excel(writer, sheet_name='Metadata', index=False)
    final_all_categories_s9.to_excel(writer, sheet_name='S9 - Category', index=False)
    final_all_categories_locale_combined_s10.to_excel(writer, sheet_name='S10 - Category - Locale', index=False)
            
# # Load the updated Excel file into a DataFrame
# comparison_s9 = pd.read_excel(output_file_path, sheet_name='S9 - Category')
# print(len(comparison_s9))
   
# comparison_s10 = pd.read_excel(output_file_path, sheet_name='S10 - Category - Locale')
# print(len(comparison_s10))

############################
## Backlog
############################
print('## Write - backlog ##')
output_file_path = os.path.join(output_folder, 'X_Backlog.xlsx')

desired_columns = ['Category', 'unique_count_GTIN', 'unique_count_VendorNumberSAP',
                   'unique_count_ArticleLongName', 'Level', 'Reason']

final_backlog_df = final_backlog_df[desired_columns]

with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
    final_backlog_df.to_excel(writer, sheet_name='Backlog', index=False)


############################
## Output in workflow
############################

metadata_original = pd.DataFrame({
    'Sheet No': ['S1', 'S2', 'S3', 'S4', 'S5', 'S6', 'S7', 'S8', 'S9', 'S10', 
                 'S11', 'S12', 'S13', 'S14', 'S14.1', 'S15', 'S16', 'S17', 
                 'S18', 'S19', 'S20', 'S21', 'S22', 'S23', 'S24', 'S25', 'S26', 'S27'],
    'DataModel Type Name': ['Organization', 'Hierarchy', 'Container', 'Container - Locale', 'Entity Type', 
                            'Relationship Type', 'Attribute Model', 'Attribute Model - Localized Values', 
                            'Category', 'Category - Localized', 'Container - Entity Type', 'Entity Type - Attribute', 
                            'Container - Entity Type - Attribute', 'Category - Attribute', 
                            'Category - Attribute - Localization', 'Relationship Type - Entity Type', 
                            'Relationship Type - Entity Type - Cardinality', 'Container - RelationshipType - Entity Type', 
                            'Container - RelationshipType - Entity Type - Cardinality', 'Relationship Type - Attribute', 
                            'Container - RelationshipType - Attribute', 'Security Role', 'Security User', 
                            'Lookup Model', 'Word List', 'Word Element', 'Entity Variant Definition', 
                            'Entity Variant Definition Mapping'],
    'Physical Sheet Name': ['S1 - Organization', 'S2 - Hierarchy', 'S3 - Container', 'S4 - Container - Locale', 
                            'S5 - Entity Type', 'S6 - Relationship Type', 'S7 - Attribute', 'S8 - Attribute - Locale', 
                            'S9 - Category', 'S10 - Category - Locale', 'S11 - CON - ET', 'S12 - ET - ATTR', 
                            'S13 - CON - ET - ATTR', 'S14 - CAT - ATTR', 'S14.1 - CAT - ATTR - Locale', 
                            'S15 - RT - ET', 'S16 - RT - ET - CARD', 'S17 - CON - RT - ET', 
                            'S18 - CON - RT - ET - CARD', 'S19 - RT - ATTR', 'S20 - CON - RT - ATTR', 
                            'S21 - Security Role', 'S22 - Security User', 'S23 - Lookup Model', 
                            'S24 - Word List', 'S25 - Word Element', 'S26 - Entity Variant Definition', 
                            'S27 - EVD Mapping']
})

def metadata_lookupvalues(dataframe):
    metadata_sheet = {
        'LookupTableName': [item['sheet_name'] for item in dataframe],
        'SheetName': [item['sheet_name'] for item in dataframe],
        'Load Lookup?': ['Yes'] * len(dataframe)
    }
    metadata_sheet_df = pd.DataFrame(metadata_sheet)

    return metadata_sheet_df

def metadata_general(physical_sheet_names):
    metadata_filtered = metadata_original[metadata_original['Physical Sheet Name'].isin(physical_sheet_names)].copy()
    metadata_filtered['Load Lookup?'] = 'YES'
    
    return metadata_filtered[['Sheet No', 'DataModel Type Name', 'Physical Sheet Name', 'Load Lookup?']]



print('## Write - Additions and deletion - S9 & S10 ##')
# Define a list of DataFrames to iterate over
dataframes_s9 = {
    'new_segment_categories_S9AndS10': new_segment_categories_s9,
    'new_family_categories_S9AndS10': new_family_categories_s9,
    'new_class_categories_S9AndS10': new_class_categories_s9,
    'new_brick_categories_S9AndS10': new_brick_categories_s9,
    'delete_brick_categories_S9_autotriggers10': delete_brick_categories_s9,
    'delete_class_categories_S9_autotriggers10': delete_class_categories_s9,
    'delete_family_categories_S9_autotriggers10': delete_family_categories_s9,
    'delete_segment_categories_S9_autotriggers10': delete_segment_categories_s9,
}

dataframes_s10 = {
    'new_segment_categories_S9AndS10': new_segment_categories_s10,
    'new_family_categories_S9AndS10': new_family_categories_s10,
    'new_class_categories_S9AndS10': new_class_categories_s10,
    'new_brick_categories_S9AndS10': new_brick_categories_s10,
}

# Define counters for numbering the files
new_file_counter = 1
delete_file_counter = 17

# Helper function to extract metadata information
def extract_metadata(sheet_name):
    parts = sheet_name.split('-')
    sheet_no = parts[0].strip()
    data_model_type_name = parts[1].strip() if len(parts) > 1 else ''
    return sheet_no, data_model_type_name

# Save each DataFrame to a separate Excel file with a prefixed number
for name, dataframe in dataframes_s9.items():
    if name.startswith('new_'):
        output_file_path = os.path.join(output_folder, f"{new_file_counter}_{name}.xlsx")
        new_file_counter += 1
    elif name.startswith('delete_'):
        output_file_path = os.path.join(output_folder, f"{delete_file_counter}_{name}.xlsx")
        delete_file_counter += 1
    else:
        continue  # Skip if the name does not match the expected prefixes

    with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:


        if name in dataframes_s10:
            sheet_name = ['S9 - Category', 'S10 - Category - Locale']
        else:
            sheet_name = ['S9 - Category']


        # Write Metadata sheet first
        metadata_df = metadata_general(sheet_name)
        metadata_df.to_excel(writer, sheet_name='Metadata', index=False)

        # Write S9 sheet
        dataframe.to_excel(writer, sheet_name='S9 - Category', index=False)

        # Write S10 sheet if it exists
        if name in dataframes_s10:
            dataframes_s10[name].to_excel(writer, sheet_name='S10 - Category - Locale', index=False)

print('## Write - Changes - S9 ##')
# Combine change_hierarchy DataFrames into one DataFrame
change_hierarchy_combined_df = pd.concat([change_hierarchy_bricks_s9, change_hierarchy_classes_s9, change_hierarchy_families_s9], ignore_index=True)

metadata_df = metadata_general(['S9 - Category'])

# Save the combined change_hierarchy DataFrame to a single Excel file
output_file_path = os.path.join(output_folder, '5_REQUIRES_DATA_MIGRATION_change_hierarchy_combined.xlsx')
with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
    metadata_df.to_excel(writer, sheet_name='Metadata', index=False)
    change_hierarchy_combined_df.to_excel(writer, sheet_name='S9 - Category', index=False)

###################
## Write output
###################
print('\n  ### Output categories ###')

output_file_path_attributes = os.path.join(output_folder, 'GS1_vs_Datamodel_Comparison_Attributes.xlsx')
output_file_path_lookupdata = os.path.join(output_folder, 'LookupData.xlsx')

############################
## Combined output
############################

metadata_df = metadata_general(['S7 - Attribute','S8 - Attribute - Locale','S23 - Lookup Model'])

# Write S7, S8 and S23
with pd.ExcelWriter(output_file_path_attributes, engine='openpyxl') as writer:
    metadata_df.to_excel(writer, sheet_name='Metadata', index=False)
    print("## Write - S7 total ##")
    final_s7_df.to_excel(writer, sheet_name='S7 - Attribute', index=False)
    print("## Write - S8 total ##")
    final_s8_df.to_excel(writer, sheet_name='S8 - Attribute - Locale', index=False)
    print("## Write - S23 total ##")
    maxeda_s23_total_df.to_excel(writer, sheet_name='S23 - Lookup Model', index=False)   


# # Write lookup tabel values
# with pd.ExcelWriter(output_file_path_lookupdata, engine='openpyxl') as writer:
#     # print("## LookupData ##")

#     # Create metadata data frame
#     metadata_lookupvalues_total = metadata_lookupvalues(final_lookupdata_df)
#     # Write the metadata DataFrame as the first sheet named 'Metadata'
#     metadata_lookupvalues_total.to_excel(writer, sheet_name='Metadata', index=False)

#     # Write each DataFrame to its respective sheet
#     for item in tqdm(final_lookupdata_df, desc="## Write - LookupData values - total ##"):
#         # Write DataFrame to a sheet named after the original sheet_name
#         item['df'].to_excel(writer, sheet_name=item['sheet_name'], index=False)

############################
## Output in workflow
############################


with pd.ExcelWriter(os.path.join(output_folder, '6_Add_LookupData_Tables_S23.xlsx'), engine='openpyxl') as writer:
    print("## Write - Additions - S23 ##")

    # Create metadata data frame
    metadata_s23_add = metadata_general(['S23 - Lookup Model'])   
    # Write the metadata DataFrame as the first sheet named 'Metadata'
    metadata_s23_add.to_excel(writer, sheet_name='Metadata', index=False)

    all_additions_attributes_s23_df.to_excel(writer, sheet_name='S23 - Lookup Model', index=False)

with pd.ExcelWriter(os.path.join(output_folder, '7_Add_LookupData_Values.xlsx'), engine='openpyxl') as writer:
    # print("## Write - Additions - LookupData values ##")
    
    # # Filter the list for items with 'filename' == 'NO FILE: addition'
    # lookupvalues_add_df = [item for item in final_add_lookup_values if item['filename'] == 'NO FILE: addition']

    # Create metadata data frame
    metadata_lookupvalues_add = metadata_lookupvalues(final_add_lookup_values)   
    # Write the metadata DataFrame as the first sheet named 'Metadata'
    metadata_lookupvalues_add.to_excel(writer, sheet_name='Metadata', index=False)

    # Write each DataFrame to its respective sheet if 'filename' is 'NO FILE: addition'
    for item in tqdm(final_add_lookup_values, desc="## Write - Additions - LookupData values ##"):
        # Write DataFrame to a sheet named after the original sheet_name
        item['df'].to_excel(writer, sheet_name=item['sheet_name'], index=False)

all_additions_attributes_s7_df = all_additions_attributes_s7_df[columns_s7]
changes_s7_df = changes_s7_df[columns_s7]

with pd.ExcelWriter(os.path.join(output_folder, '8_Add_Attributes_S7andS8.xlsx'), engine='openpyxl') as writer:
    print("## Write - Additions - S7 & S8 + Changes S7 ##")

    # Create metadata data frame
    metadata_s7_add = metadata_general(['S7 - Attribute', 'S8 - Attribute - Locale'])   
    # Write the metadata DataFrame as the first sheet named 'Metadata'
    metadata_s7_add.to_excel(writer, sheet_name='Metadata', index=False)

    all_additions_attributes_s7_df.to_excel(writer, sheet_name='S7 - Attribute', index=False)
    all_additions_attributes_s8_df.to_excel(writer, sheet_name='S8 - Attribute - Locale', index=False)

with pd.ExcelWriter(os.path.join(output_folder, '9_Add_CDB_MDM_Mapping.xlsx'), engine='openpyxl') as writer:
    print("## Write - Additions - CDB_MDM_Mapping ##")

    # Create metadata data frame
    metadata_sheet = pd.DataFrame({
        'LookupTableName': ['CDB_MDMAttributeMapping', 'CDB_GS1Attributes'],
        'SheetName': ['CDB_MDMAttributeMapping', 'CDB_GS1Attributes'],
        'Load Lookup?': ['Yes', 'Yes']
    })

    # Write the metadata DataFrame as the first sheet named 'Metadata'
    metadata_sheet.to_excel(writer, sheet_name='Metadata', index=False)
    cdb_mdm_add_df.to_excel(writer, sheet_name='CDB_MDMAttributeMapping', index=False)
    cdb_gs1_add_df.to_excel(writer, sheet_name='CDB_GS1Attributes', index=False)
    
    
with pd.ExcelWriter(os.path.join(output_folder, '10_Delete_CDB_MDM_Mapping.xlsx'), engine='openpyxl') as writer:
    print("## Write - Additions - CDB_MDM_Mapping ##")

    # Create metadata data frame
    metadata_sheet = pd.DataFrame({
        'LookupTableName': ['CDB_MDMAttributeMapping', 'CDB_GS1Attributes'],
        'SheetName': ['CDB_MDMAttributeMapping', 'CDB_GS1Attributes'],
        'Load Lookup?': ['Yes', 'Yes']
    })

    # Write the metadata DataFrame as the first sheet named 'Metadata'
    metadata_sheet.to_excel(writer, sheet_name='Metadata', index=False)
    cdb_mdm_delete_df.to_excel(writer, sheet_name='CDB_MDMAttributeMapping', index=False)
    cdb_gs1_delete_df.to_excel(writer, sheet_name='CDB_GS1Attributes', index=False)
    

with pd.ExcelWriter(os.path.join(output_folder, '11_REQUIRES_DATA_MIGRATION_Change_Attributes_S7.xlsx'), engine='openpyxl') as writer:
    print("## Write - Additions - S7 & S8 + Changes S7 ##")

    # Create metadata data frame
    metadata_s7_add = metadata_general(['S7 - Attribute'])   
    # Write the metadata DataFrame as the first sheet named 'Metadata'
    metadata_s7_add.to_excel(writer, sheet_name='Metadata', index=False)

    changes_s7_df.to_excel(writer, sheet_name='S7 - Attribute', index=False)

with pd.ExcelWriter(os.path.join(output_folder, '12_Add_Brick_Attribute_Combinations_S14.xlsx'), engine='openpyxl') as writer:
    print("## Write - Additions - S14 ##")
    
    # Create metadata data frame
    metadata_s14_add = metadata_general(['S14 - CAT - ATTR'])   
    # Write the metadata DataFrame as the first sheet named 'Metadata'
    metadata_s14_add.to_excel(writer, sheet_name='Metadata', index=False)
    
    add_S14_df.to_excel(writer, sheet_name='S14 - CAT - ATTR', index=False)
   

    ############################
    ## Manual lookup table value deletion
    ############################
print('## Write - manual - Lookup table value deletion ##')
output_file_path = os.path.join(output_folder, '13_MANUAL_PROCESSING_Delete_LookupData_Values.xlsx')

with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
    final_delete_lookup_values.to_excel(writer, sheet_name='Lookup table value deletion', index=False)

# # Write deletions of lookup data values
# with pd.ExcelWriter(os.path.join(output_folder, '9_Delete_LookupData_Values.xlsx'), engine='openpyxl') as writer:
#     # print("## Write - Deletions - LookupData delete values ##")
 
#     # Create metadata data frame
#     metadata_lookupvalues_delete = metadata_lookupvalues(delete_lookupdata_df)   
#     # Write the metadata DataFrame as the first sheet named 'Metadata'
#     metadata_lookupvalues_delete.to_excel(writer, sheet_name='Metadata', index=False)

#     # Write each DataFrame to its respective sheet
#     for item in tqdm(delete_lookupdata_df, desc="## Write - Deletions - LookupData delete values ##"):
#         # Write DataFrame to a sheet named after the original sheet_name
#         item['df'].to_excel(writer, sheet_name=item['sheet_name'], index=False)


with pd.ExcelWriter(os.path.join(output_folder, '14_Delete_Brick_Attribute_Combinations_S14.xlsx'), engine='openpyxl') as writer:
    print("## Write - Deletions - S14 ##")
    
    # Create metadata data frame
    metadata_s14_delete = metadata_general(['S14 - CAT - ATTR'])   
    # Write the metadata DataFrame as the first sheet named 'Metadata'
    metadata_s14_delete.to_excel(writer, sheet_name='Metadata', index=False)
    
    maxeda_s14_delete_df.to_excel(writer, sheet_name='S14 - CAT - ATTR', index=False)


with pd.ExcelWriter(os.path.join(output_folder, '15_Delete_LookupData_Tables_S23.xlsx'), engine='openpyxl') as writer:
    print("## Write - Deletions - S23 ##")
    
    # Create metadata data frame
    metadata_s23_delete = metadata_general(['S23 - Lookup Model'])   
    # Write the metadata DataFrame as the first sheet named 'Metadata'
    metadata_s23_delete.to_excel(writer, sheet_name='Metadata', index=False)
    
    maxeda_s23_delete_df.to_excel(writer, sheet_name='S23 - Lookup Model', index=False)


# with pd.ExcelWriter(os.path.join(output_folder, '3_Delete_Attributes_S8.xlsx'), engine='openpyxl') as writer:
#     print("## 1.3 - Deletions - Attributes S8 ##")

#     # Create metadata data frame
#     metadata_s8_delete = metadata_general(['S8 - Attribute - Locale'])   
#     # Write the metadata DataFrame as the first sheet named 'Metadata'
#     metadata_s8_delete.to_excel(writer, sheet_name='Metadata', index=False)

#     delete_attributes_s8_df.to_excel(writer, sheet_name='S8 - Attribute - Locale', index=False)

with pd.ExcelWriter(os.path.join(output_folder, '16_Delete_Attributes_S7_autotriggersS8.xlsx'), engine='openpyxl') as writer:
    print("## Write - Deletions - Attributes S7 autotrigger S8 ##")

    # Create metadata data frame
    metadata_s7_delete = metadata_general(['S7 - Attribute'])   
    # Write the metadata DataFrame as the first sheet named 'Metadata'
    metadata_s7_delete.to_excel(writer, sheet_name='Metadata', index=False)

    delete_attributes_s7_df.to_excel(writer, sheet_name='S7 - Attribute', index=False)


# End timing
end_time = time.time()

# Calculate elapsed time in seconds
elapsed_time = end_time - start_time

# Calculate elapsed time in seconds
elapsed_time = end_time - start_time

# Measure CPU and memory after the code block
end_cpu_usage, end_memory_usage = get_cpu_memory_usage()

# Calculate CPU and memory usage differences
cpu_usage_diff = end_cpu_usage - start_cpu_usage
memory_usage_diff = end_memory_usage - start_memory_usage

# Convert to minutes and seconds
minutes = int(elapsed_time // 60)
seconds = elapsed_time % 60

print(f"Elapsed time: {minutes} minutes and {seconds:.2f} seconds")
print(f"CPU usage: {cpu_usage_diff:.2f}%")
print(f"Memory usage: {memory_usage_diff:.2f} MB")



# ####################
# ## Testing
# ####################            
# # Load the updated Excel file into a DataFrame to confirm it saved correctly
# loaded_attributes_s7_df = pd.read_excel(output_file_path_attributes, sheet_name='S7 - Attribute')
# loaded_attributes_s8_df = pd.read_excel(output_file_path_attributes, sheet_name='S8 - Attribute - Locale')
# loaded_lookupdata_df_metadata = pd.read_excel(output_file_path_lookupdata, sheet_name='Metadata')

# Expected_s7_additions = 58
# Expected_s7_deletions = 287
# Expected_s7_changes = 710
# loaded_attributes_s7 = 1191 # vs. calc >1113 (2* additions - Category & Common OneWS, 1.X * delete - OFTEN MISSING attribute code for Common OneWS, Changes - 1.X * change, some Attribute code have multiple rows)


# Expected_s8_additions = Expected_s7_additions * 4 * 2 # languages + OneWS
# Expected_attribute_delete_s7_MaxedaIDs_set = 329 # vs calc 287 * (1 + X). ? Because only 2 Common OneWS have attribute code incorporated, some ID's have multiple rows
# Expected_s8_deletions = Expected_attribute_delete_s7_MaxedaIDs_set * 4 # languages
# loaded_attributes_s8 = Expected_s8_additions + Expected_s8_deletions # vs. calc 1780  (Expected_s8_additions + (4 * delete considering direct ID'))

# expected_items_metadata_lookupdata = 373
# loaded_sheets_lookupdata = 374

# assert len(attribute_add_s7_GS1ID_set) == Expected_s7_additions, f"Expected {Expected_s7_additions} additions, got {len(attribute_add_s7_GS1ID_set)}"
# assert len(attribute_delete_s7_GS1ID_set) == Expected_s7_deletions, f"Expected {Expected_s7_deletions} deletions, got {len(attribute_delete_s7_GS1ID_set)}"
# assert len(change_set) == Expected_s7_changes, f"Expected {Expected_s7_changes} changes, got {len(change_set)}"
# assert len(loaded_attributes_s7_df) == loaded_attributes_s7, f"Expected {loaded_attributes_s7} total entries in S7, got {len(loaded_attributes_s7_df)}"

# assert len(all_additions_attributes_s8_df) == Expected_s8_additions, f"Expected {Expected_s8_additions} additions locale in S8, got {len(all_additions_attributes_s8_df)}"
# assert len(attribute_delete_s7_MaxedaIDs_set) == Expected_attribute_delete_s7_MaxedaIDs_set , f"Expected {Expected_attribute_delete_s7_MaxedaIDs_set} in ID-set, got {len(attribute_delete_s7_MaxedaIDs_set)}"
# assert len(loaded_attributes_s8_df) == loaded_attributes_s8, f"Expected {loaded_attributes_s8} total entries in S8, got {len(loaded_attributes_s8_df)}"

# assert len(loaded_lookupdata_df_metadata) == expected_items_metadata_lookupdata, f"Metadata sheet should contain exactly {expected_items_metadata_lookupdata} items, found {len(loaded_lookupdata_df_metadata)} items."
# actual_sheets_count = len(pd.ExcelFile(output_file_path_lookupdata).sheet_names)
# assert actual_sheets_count == loaded_sheets_lookupdata, f"Excel file should contain {loaded_sheets_lookupdata} sheets, found {actual_sheets_count} sheets."
# # Check if the third sheet is named 'TOclothingforpets'
# third_sheet_name = pd.ExcelFile(output_file_path_lookupdata).sheet_names[2]  # Indexing starts from 0
# assert third_sheet_name == 'TOclothingforpets', f"The third sheet is named {third_sheet_name}, not 'TOclothingforpets'."
# # Load the third sheet
# third_sheet_df = pd.read_excel(output_file_path_lookupdata, sheet_name=third_sheet_name)
# # Assert that this sheet has 28 items (rows)
# assert len(third_sheet_df) == 28, f"The sheet 'TOclothingforpets' has {len(third_sheet_df)} items, not 28."


# print("All tests passed successfully.")
